import os
from django.shortcuts import get_object_or_404, render
from rest_framework.views import APIView, Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from .serializers import *
import http.client
import json
import random
import xmltodict
from datetime import datetime
import string
import requests
from rest_framework.decorators import api_view, permission_classes
from dotenv import load_dotenv
import xml.sax.saxutils as saxutils 
from django.contrib.auth import authenticate
from openpyxl import load_workbook
from datetime import date, datetime, time

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '../.env'))
base_url = os.environ.get('Base_url')
loginid = os.environ.get('LoginId')
membercode = os.environ.get('MemberCode')
password = os.environ.get('Password')
passkey = os.environ.get('PassKey')

escaped_password = saxutils.escape(password)
def get_password():
    try:
        soap_message = f"""
        <soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope"
                    xmlns:ns="http://bsestarmfdemo.bseindia.com/2016/01/">
            <soap:Header xmlns:wsa="http://www.w3.org/2005/08/addressing">
                <wsa:Action>http://bsestarmfdemo.bseindia.com/2016/01/IMFUploadService/getPassword</wsa:Action>
                <wsa:To>https://bsestarmfdemo.bseindia.com/MFUploadService/MFUploadService.svc/Secure</wsa:To>
            </soap:Header>
            <soap:Body>
                <ns:getPassword>
                    <ns:UserId>{loginid}</ns:UserId>
                    <ns:MemberId>{membercode}</ns:MemberId>
                    <ns:Password>{escaped_password}</ns:Password>
                    <ns:PassKey>{passkey}</ns:PassKey>
                </ns:getPassword>
            </soap:Body>
        </soap:Envelope>
        """

        conn = http.client.HTTPSConnection(base_url)
        headers = {
            'Content-Type': 'application/soap+xml; charset=utf-8',
            'SOAPAction': 'http://bsestarmfdemo.bseindia.com/2016/01/IMFUploadService/getPassword'
        }

        conn.request("POST", "/MFUploadService/MFUploadService.svc/Secure", soap_message, headers)
        res = conn.getresponse()
        raw_response = res.read().decode()

        print("\n\n==== RAW PASSWORD RESPONSE ====")
        print(raw_response)
        print("================================\n\n")

        if res.status != 200:
            raise Exception(f"Failed to retrieve password, Status: {res.status}")

        try:
            api_data_dict = xmltodict.parse(raw_response)
            password_result = api_data_dict.get('s:Envelope', {}).get('s:Body', {}).get('getPasswordResponse', {}).get('getPasswordResult', '')
        except Exception as e:
            raise Exception(f"XML Parsing Error: {str(e)}")

        if not password_result:
            raise Exception("Password not found in response")

        if "|" in password_result:
            _, password = password_result.split("|", 1)
        else:
            password = password_result

        return password.strip()
    except Exception as e:
        return {"status": False, "message": str(e)}

def generate_unique_code():
    prefix = "WR"
    random_number = random.randint(1000000000, 9999999999)
    return f"{random_number}"

class TestAPiView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        return Response({"message": "It is working"}, status=status.HTTP_200_OK)

class SignupAPIView(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        response_data = {}

        user_id = request.GET.get('id')

        if user_id:
            try:
                user = User.objects.get(id=user_id)
                serializer = UserRegistrationSerializer(user)
                response_data['status'] = True
                response_data['message'] = 'User details fetched successfully'
                response_data['data'] = serializer.data
                return Response(response_data, status=status.HTTP_200_OK)
            except User.DoesNotExist:
                response_data['status'] = False
                response_data['message'] = 'User not found'
                return Response(response_data, status=status.HTTP_404_NOT_FOUND)
        else:
            users = User.objects.all()
            serializer = UserRegistrationSerializer(users, many=True)
            response_data['status'] = True
            response_data['message'] = 'User details fetched successfully'
            response_data['data'] = serializer.data
            return Response(response_data, status=status.HTTP_200_OK)

    def post(self,request):
        response_data = {}
        data = request.data

        if hasattr(data, '_mutable'):
            data._mutable = True

        email = data.get('email')
        if User.objects.filter(username=email).exists():
            response_data['status'] = False
            response_data['message'] = 'A user with this email already exists.'
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        data['username'] = data.get('phone')
        serializer = UserRegistrationSerializer(data=data)
        
        if serializer.is_valid():
            user = serializer.save()
            user.set_password(data.get("password"))
            user.save()
            response_data['status']= True
            response_data['message']= 'User register successfully'
            response_data['data']= serializer.data
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            response_data['status']= False
            response_data ['message']= 'User register failed'
            response_data['errors']= serializer.errors
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
def generate_unique_otp():
        while True:
            otp = ''.join(random.choices(string.digits, k=5))
            if not User.objects.filter(otp=otp).exists():
                return otp
            
class SendOtpView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        phone = request.data.get('phone')
        name = request.data.get('name')
    
        if not phone:
            return Response({
                'success': False,
                'message': 'Phone number is required.',
                'error': {'phone': ['This field is required.']}
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.filter(phone=phone).first()

            if user:
                if Agent.objects.filter(user=user).exists():
                    return Response({
                        'success': False,
                        'message': 'You are already registered as an agent.',
                        'error': {'agent': 'Agent already exists for this phone number.'}
                    }, status=status.HTTP_400_BAD_REQUEST)

                otp = generate_unique_otp()
                user.otp = otp
                user.save()
            else:
                otp = generate_unique_otp()
                user = User.objects.create(username=phone, phone=phone, name=name, otp=otp)

            sms_text = str(otp)
            api_url = (
                f"https://www.proactivesms.in/sendsms.jsp?"
                f"user=singhsft&password=b03ec2c66eXX&senderid=SINGSF&"
                f"mobiles={phone}&sms={sms_text}&tempid=1707171394331116655"
            )
            sms_response = requests.get(api_url)

            if sms_response.status_code == 200:
                user_data = UserRegistrationSerializer(user).data
                return Response({
                    'success': True,
                    'message': 'OTP sent successfully.',
                    'data': user_data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'message': 'Failed to send OTP.',
                    'error': {'sms': 'SMS API failed to send the message.'}
                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'success': False,
                'message': 'An error occurred while sending OTP.',
                'error': {'exception': str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# from .models import User, Agent

# @api_view(['POST'])
# @permission_classes([AllowAny])
# def verify_otp(request):
#     response_data = {}

#     email = request.GET.get('email')  # use email instead of phone
#     otp = request.data.get('otp')

#     if not email or not otp:
#         response_data['status'] = False
#         response_data['message'] = 'Email and OTP are required'
#         response_data['errors'] = {
#             'email': ['This field is required'],
#             'otp': ['This field is required']
#         }
#         return Response(response_data, status=400)

#     try:
#         user = User.objects.filter(email=email).first()
#         if not user:
#             response_data['status'] = False
#             response_data['message'] = 'No user found with this email'
#             return Response(response_data, status=404)

#         if user.otp == otp:
#             user.otp = None
#             user.save(update_fields=['otp'])

#             # Clear agent OTP if exists
#             agent = Agent.objects.get(user=user)
#             if agent:
#                 agent.otp = None
#                 agent.save(update_fields=['otp'])

#             response_data['status'] = True
#             response_data['message'] = 'OTP verified successfully'
#             response_data['data'] = {
#                 'user_id': user.id,
#                 'email': user.email
#             }
#             return Response(response_data, status=200)
#         else:
#             response_data['status'] = False
#             response_data['message'] = 'Wrong or invalid OTP'
#             response_data['errors'] = {'otp': ['Invalid OTP']}
#             return Response(response_data, status=400)

#     except User.DoesNotExist:
#         response_data['status'] = False
#         response_data['message'] = 'User not found'
#         response_data['errors'] = {'phone': ['User not registered']}
#         return Response(response_data, status=status.HTTP_404_NOT_FOUND)




@api_view(('POST',))
@permission_classes([AllowAny])       
def verify_otp(request):
    response_data = {}
    email = request.GET.get('email')
    print(email)
    phone = request.GET.get('phone')
    otp = request.data.get('otp')

    if not email or not otp:
        response_data['status'] = False
        response_data['message'] = 'email and OTP are required'
        response_data['errors'] = {'email': ['This field is required'], 'otp': ['This field is required']}
        return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

    try:
        # user = User.objects.get(phone=phone)
        user = User.objects.get(email=email)
        print(user)
        if user.otp == otp:
            user.otp = None
            user.save()

            try:
                agent = Agent.objects.get(user=user)
                agent.otp = None
                agent.save()
            except Agent.DoesNotExist:
                pass

            response_data['status'] = True
            response_data['message'] = 'OTP verified successfully'
            response_data['data'] = {
                'user_id': user.id,
                'phone': user.email
            }
            return Response(response_data, status=status.HTTP_200_OK)
        else:
            response_data['status'] = False
            response_data['message'] = 'Wrong or invalid OTP'
            response_data['errors'] = {'otp': ['Invalid OTP']}
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

    except User.DoesNotExist:
        response_data['status'] = False
        response_data['message'] = 'User not found'
        response_data['errors'] = {'phone': ['User not registered']}
        return Response(response_data, status=status.HTTP_404_NOT_FOUND)
   
class loginAPIView(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        users = User.objects.all()
        serializer = UserRegistrationSerializer(users, many=True)
        return Response(serializer.data)
    
    # def post(self, request):
    #     phone = request.data.get('phone')
    #     password = request.data.get('password')
    #     print(phone)

    #     if not phone:
    #         return Response({
    #             'status': False,
    #             'message': 'Phone number is required',
    #             'data': []
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     if not password:
    #         return Response({
    #             'status': False,
    #             'message': 'Password is required',
    #             'data': []
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         user_obj = User.objects.get(username=phone)
    #     except User.DoesNotExist:
    #         return Response({
    #             'status': False,
    #             'message': 'Phone number not registered',
    #             'data': []
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     user = authenticate(username=phone, password=password)
    #     print("User Obj:", user_obj.username)
    #     print("Password is set:", user_obj.password)
    #     if user is None:
    #         return Response({
    #             'status': False,
    #             'message': 'Incorrect password',
    #             'data': []
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     serializer = UserRegistrationSerializer(user)
    #     return Response({
    #         'status': True,
    #         'message': 'User logged in successfully',
    #         'data': serializer.data
    #     }, status=status.HTTP_200_OK)
     #pehele comment the below code   
    def post(self, request):
        try:
            response_data = {}
            data = request.data.copy()
            data._mutable = True
            phone = data.get('phone')

            if not phone:
                response_data['status'] = False
                response_data['message'] = "Phone number is required"
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            try:
                user = User.objects.get(phone=phone)
            except User.DoesNotExist:
                response_data['status'] = False
                response_data['message'] = "User with this phone number does not exist"
                return Response(response_data, status=status.HTTP_404_NOT_FOUND)

            otp = generate_unique_otp()
            user.otp = otp
            user.save()

            phone_number = user.phone
            name = user.name
            otp_code = otp

            api_url = (
                f"https://www.proactivesms.in/sendsms.jsp?"
                f"user=singhsft&password=b03ec2c66eXX&senderid=SINGSF&mobiles={phone_number}"
                f"&sms=Dear+{name}+Please+enter+the+OTP+code+{otp_code}+to+confirm+your+mobile+number+at+Singhsoft"
                f"&tempid=1707171394331116655"
            )
            response = requests.get(api_url)

            if response.status_code == 200:
                serializer = UserRegistrationSerializer(user)
                response_data['status'] = True
                response_data['message'] = 'OTP sent successfully to your phone number'
                response_data['data'] = serializer.data
                return Response(response_data, status=status.HTTP_200_OK)
            else:
                response_data['status'] = False
                response_data['message'] = 'Error sending OTP to phone number'
                response_data['errors'] = 'SMS API error'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            response_data['status'] = False
            response_data['message'] = "An error occurred while processing your request."
            response_data['errors'] = str(e)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
from django.core.mail import send_mail
from django.conf import settings       
class ResendOTPAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        email = request.data.get("email")  # changed from phone to email

        # Check if email is provided
        if not email:
            response_data["status"] = False
            response_data["message"] = "Email is required."
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(email=email)

            if not user.email:
                response_data["status"] = False
                response_data["message"] = "User does not have an email address set."
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            # Generate new OTP
            otp = generate_unique_otp()
            user.otp = otp
            user.save(update_fields=["otp"])

            # Prepare email
            name = user.name or "User"
            subject = "Your OTP Code (Resent)"
            message = f"Hello {name},\n\nYour new OTP code is: {otp}\n\nPlease do not share this code with anyone."
            email_from = settings.DEFAULT_FROM_EMAIL
            recipient_list = [email]

            # Send the email
            send_mail(subject, message, email_from, recipient_list)

            response_data["status"] = True
            response_data["message"] = "OTP resent successfully to your email."
            return Response(response_data, status=status.HTTP_200_OK)

        except User.DoesNotExist:
            response_data["status"] = False
            response_data["message"] = "User with this email does not exist."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            response_data["status"] = False
            response_data["message"] = "An error occurred while resending OTP."
            response_data["error"] = str(e)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        
            
# class ResendOTPAPIView(APIView):
#     permission_classes = [AllowAny]

#     def post(self, request):
#         response_data = {}
#         phone = request.data.get("phone")

#         # Check if phone is provided
#         if not phone:
#             response_data["status"] = False
#             response_data["message"] = "Phone number is required."
#             return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             user = User.objects.get(phone=phone)

#             if not user.phone:
#                 response_data["status"] = False
#                 response_data["message"] = "User does not have a phone number set."
#                 return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

#             otp = generate_unique_otp()
#             user.otp = otp
#             user.save(update_fields=["otp"])

#             name = user.name or "User"
#             mobile_number = user.phone
#             otp_code = otp
#             api_url = f"https://www.proactivesms.in/sendsms.jsp?user=singhsft&password=b03ec2c66eXX&senderid=SINGSF&mobiles={mobile_number}&sms=Dear+{name}+Please+enter+the+OTP+code+{otp_code}+to+confirm+your+mobile+number+at+Singhsoft&tempid=1707171394331116655"

#             sms_response = requests.get(api_url)

#             if sms_response.status_code == 200:
#                 response_data["status"] = True
#                 response_data["message"] = "OTP resent successfully."
#                 return Response(response_data, status=status.HTTP_200_OK)
#             else:
#                 response_data["status"] = False
#                 response_data["message"] = "Failed to send OTP."
#                 return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         except User.DoesNotExist:
#             response_data["status"] = False
#             response_data["message"] = "User with this phone number does not exist."
#             return Response(response_data, status=status.HTTP_404_NOT_FOUND)

#         except Exception as e:
#             response_data["status"] = False
#             response_data["message"] = "An error occurred while resending OTP."
#             response_data["error"] = str(e)
#             return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class LogoutAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data["refresh"]
            token = RefreshToken(refresh_token)
            token.blacklist()

            return Response({"detail": "Logout successful"}, status=status.HTTP_205_RESET_CONTENT)
        except Exception as e:
            return Response({"error": "Invalid token or already blacklisted"}, status=status.HTTP_400_BAD_REQUEST)

class AgentRegisterAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        response_data = {}
        agent_id = request.GET.get('agent_id')

        if not agent_id:
            response_data['status'] = False
            response_data['message'] = 'Agent ID is required'
            response_data['errors'] = {'agent_id': 'Missing agent_id in request'}
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(agent_id=agent_id)
        except Agent.DoesNotExist:
            response_data['status'] = False
            response_data['message'] = 'Agent not found'
            response_data['errors'] = {'agent_id': 'Invalid or non-existent agent_id'}
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        sips = XSIPTransaction.objects.filter(agent=agent).select_related('user')
        sip_data = XSIPSerializer(sips, many=True).data

        response_data['status'] = True
        response_data['message'] = 'SIP list fetched successfully'
        response_data['data'] = sip_data
        return Response(response_data, status=status.HTTP_200_OK)
    
    def post(self, request):
        response_data = {}

        try:
            data = request.data

            if hasattr(data, '_mutable'):
                data._mutable = True

            agent_id = request.GET.get('agent_id')
            if agent_id and Agent.objects.filter(agent_id=agent_id).exists():
                agent = Agent.objects.get(agent_id=agent_id)
                serializer = AgentSerializer(agent, data=data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'success': True,
                        'message': 'Agent updated successfully',
                        'data': serializer.data
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        'success': False,
                        'message': 'Validation failed',
                        'error': serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)

            user_id = data.get('user')
            
            if not user_id:
                response_data['success'] = False
                response_data['message'] = 'User ID is required.'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                response_data['success'] = False
                response_data['message'] = 'User with given ID does not exist.'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            new_password = data.get('password')
            if not new_password:
                response_data['success'] = False
                response_data['message'] = 'Password ID is required.'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
            if new_password:
                user.set_password(new_password)
                user.save()

            if Agent.objects.filter(user=user).exists():
                response_data['success'] = False
                response_data['message'] = 'Agent for this user already exists.'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            if 'agent_id' not in data:
                data['agent_id'] = generate_unique_code()

            data['user'] = user.id

            agent_serializer = AgentSerializer(data=data)

            if agent_serializer.is_valid():
                agent = agent_serializer.save()

                refresh = RefreshToken.for_user(user)
                access_token = {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }

                response_data['success'] = True
                response_data['message'] = 'User created successfully'
                response_data['data'] = agent_serializer.data
                response_data['data']['access_token'] = access_token
                return Response(response_data, status=status.HTTP_201_CREATED)

            else:
                response_data['success'] = False
                response_data['message'] = 'Agent validation failed'
                response_data['error'] = agent_serializer.errors
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            response_data['success'] = False
            response_data['message'] = 'An error occurred while saving the agent.'
            response_data['error'] = str(e)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class RegistrationAPIView(APIView):
    permission_classes = [AllowAny]
    def get(self, request):
        response_data = {}
        user_id = request.GET.get('id')

        if user_id:
            try:
                user = User.objects.get(pk=user_id)
            except User.DoesNotExist:
                response_data['errors'] = "User detail not found"
                return Response(response_data, status=status.HTTP_404_NOT_FOUND)

            serializer = UserRegistrationSerializer(user)
            response_data['status'] = True
            response_data['message'] = 'User details retrieved successfully'
            response_data['data'] = serializer.data
        else:
            users = User.objects.all()
            serializer = UserRegistrationSerializer(users, many=True)

            response_data['status'] = True
            response_data['message'] = 'All user details retrieved successfully'
            response_data['data'] = serializer.data

        return Response(response_data, status=status.HTTP_200_OK)
    
    def put(self, request):
        response_data = {}
        data = request.data

        if request.method == 'PUT':
            regn_type = request.data.get('regn_type', "")
            primary_holder_first_name = request.data.get('primary_holder_first_name', "")
            primary_holder_middle_name = request.data.get('primary_holder_middle_name', "")
            primary_holder_last_name = request.data.get('primary_holder_last_name', "")
            tax_status = request.data.get('tax_status', "")
            gender = request.data.get('gender', "")
            primary_holder_dob_incorporation = request.data.get('primary_holder_dob_incorporation', "")
            occupation_code = request.data.get('occupation_code', "")
            holding_nature = request.data.get('holding_nature', "")
            Second_Holder_First_Name  = request.data.get('Second_Holder_First_Name', "")
            Second_Holder_Middle_Name  = request.data.get('Second_Holder_Middle_Name', "")
            Second_Holder_Last_Name  = request.data.get('Second_Holder_Last_Name', "")
            Third_Holder_First_Name  = request.data.get('Third_Holder_First_Name', "")
            Third_Holder_Middle_Name  = request.data.get('Third_Holder_Middle_Name', "")
            Third_Holder_Last_Name  = request.data.get('Third_Holder_Last_Name', "")
            Second_Holder_DOB = request.data.get('Second_Holder_DOB', "")
            Third_Holder_DOB = request.data.get('Third_Holder_DOB', "")
            Guardian_First_Name = request.data.get('Guardian_First_Name', "")
            Guardian_Middle_Name = request.data.get('Guardian_Middle_Name', "")
            Guardian_Last_Name = request.data.get('Guardian_Last_Name', "")
            Guardian_DOB = request.data.get('Guardian_DOB', "")
            Primary_Holder_PAN_Exempt = request.data.get('Primary_Holder_PAN_Exempt', "")
            Second_Holder_PAN_Exempt = request.data.get('Second_Holder_PAN_Exempt', "")
            Third_Holder_PAN_Exempt = request.data.get('Third_Holder_PAN_Exempt', "")
            Guardian_PAN_Exempt = request.data.get('Guardian_PAN_Exempt', "")
            Primary_Holder_PAN = request.data.get('Primary_Holder_PAN', "")
            Second_Holder_PAN = request.data.get('Second_Holder_PAN', "")
            Third_Holder_PAN = request.data.get('Third_Holder_PAN', "")
            Guardian_PAN = request.data.get('Guardian_PAN', "")

            Primary_Holder_Exempt_Category = request.data.get('Primary_Holder_Exempt_Category', "")
            Second_Holder_Exempt_Category = request.data.get('Second_Holder_Exempt_Category', "")
            Third_Holder_Exempt_Category = request.data.get('Third_Holder_Exempt_Category', "")
            Guardian_Exempt_Category = request.data.get('Guardian_Exempt_Category', "")
            Client_Type = request.data.get('Client_Type', "")
            pms = request.data.get('pms', "")
            Default_DP = request.data.get('Default_DP', "")
            cdsl_dpid = request.data.get('cdsl_dpid', "")
            cdslctdid = request.data.get('cdslctdid', "")
            CMBP_Id = request.data.get('CMBP_Id', "")
            nsdl_dpid = request.data.get('nsdl_dpid', "")
            nsdlcltid = request.data.get('nsdlcltid', "")
            Account_Type_1 = request.data.get('Account_Type_1', "")
            Account_No_1 = request.data.get('Account_No_1', "")
            MICR_No_1 = request.data.get('MICR_No_1', "")
            IFSC_Code_1 = request.data.get('IFSC_Code_1', "")
            Default_Bank_Flag_1 = request.data.get('Default_Bank_Flag_1', "")
            Account_Type_2 = request.data.get('Account_Type_2', "")
            Account_No_2 = request.data.get('Account_No_2', "")
            MICR_No_2 = request.data.get('MICR_No_2', "")
            IFSC_Code_2 = request.data.get('IFSC_Code_2', "")
            Default_Bank_Flag_2 = request.data.get('Default_Bank_Flag_2', "")
            Account_Type_3 = request.data.get('Account_Type_3', "")
            Account_No_3 = request.data.get('Account_No_3', "")
            MICR_No_3 = request.data.get('MICR_No_3', "")

            IFSC_Code_3 = request.data.get('IFSC_Code_3', "")
            Default_Bank_Flag_3 = request.data.get('Default_Bank_Flag_3', "")
            Account_Type_4 = request.data.get('Account_Type_4', "")
            Account_No_4 = request.data.get('Account_No_4', "")
            MICR_No_4 = request.data.get('MICR_No_4', "")
            IFSC_Code_4 = request.data.get('IFSC_Code_4', "")
            Default_Bank_Flag_4 = request.data.get('Default_Bank_Flag_4', "")
            Account_Type_5 = request.data.get('Account_Type_5', "")
            Account_No_5 = request.data.get('Account_No_5', "")
            MICR_No_5 = request.data.get('MICR_No_5', "")
            IFSC_Code_5 = request.data.get('IFSC_Code_5', "")
            Default_Bank_Flag_5 = request.data.get('Default_Bank_Flag_5', "")
            Cheque_Name = request.data.get('Cheque_Name', "")
            Div_Pay_Mode = request.data.get('Div_Pay_Mode', "")
            Address_1 = request.data.get('Address_1', "")
            Address_2 = request.data.get('Address_2', "")
            Address_3 = request.data.get('Address_3', "")
            City = request.data.get('City', "")
            State = request.data.get('State', "")
            Pincode = request.data.get('Pincode', "")
            Country = request.data.get('Country', "")
            Resi_Phone = request.data.get('Resi_Phone', "")
            Resi_Fax = request.data.get('Resi_Fax', "")
            Office_Phone = request.data.get('Office_Phone', "")
            Office_Fax = request.data.get('Office_Fax', "")
            Email = request.data.get('Email', "")
            Communication_Mode = request.data.get('Communication_Mode', "")
            Foreign_Address_1 = request.data.get('Foreign_Address_1', "")

            Foreign_Address_2 = request.data.get('Foreign_Address_2', "")
            Foreign_Address_3 = request.data.get('Foreign_Address_3', "")
            Foreign_Address_City = request.data.get('Foreign_Address_City', "")
            Foreign_Address_Pincode = request.data.get('Foreign_Address_Pincode', "")
            Foreign_Address_State = request.data.get('Foreign_Address_State', "")
            Foreign_Address_Country = request.data.get('Foreign_Address_Country', "")
            Foreign_Address_Resi_Phone = request.data.get('Foreign_Address_Resi_Phone', "")
            Foreign_Address_Fax = request.data.get('Foreign_Address_Fax', "")
            Foreign_Address_Off_Phone = request.data.get('Foreign_Address_Off_Phone', "")
            Foreign_Address_Off_Fax = request.data.get('Foreign_Address_Off_Fax', "")
            Indian_Mobile_No = request.data.get('Indian_Mobile_No', "")
            Nominee_1_Name = request.data.get('Nominee_1_Name', "")
            Nominee_1_Relationship = request.data.get('Nominee_1_Relationship', "")
            Nominee_1_Applicable = request.data.get('Nominee_1_Applicable', "")
            Nominee_1_Minor_Flag = request.data.get('Nominee_1_Minor_Flag', "")
            Nominee_1_DOB = request.data.get('Nominee_1_DOB', "")
            Nominee_1_Guardian = request.data.get('Nominee_1_Guardian', "")
            Nominee_2_Name = request.data.get('Nominee_2_Name', "")
            Nominee_2_Relationship = request.data.get('Nominee_2_Relationship', "")
            Nominee_2_Applicable = request.data.get('Nominee_2_Applicable', "")
            Nominee_2_DOB = request.data.get('Nominee_2_DOB', "")
            Nominee_2_Minor_Flag = request.data.get('Nominee_2_Minor_Flag', "")
            Nominee_2_Guardian = request.data.get('Nominee_2_Guardian', "")
            Nominee_3_Name = request.data.get('Nominee_3_Name', "")

            Nominee_3_Relationship = request.data.get('Nominee_3_Relationship', "")
            Nominee_3_Applicable = request.data.get('Nominee_3_Applicable', "")
            Nominee_3_DOB = request.data.get('Nominee_3_DOB', "")
            Nominee_3_Minor_Flag = request.data.get('Nominee_3_Minor_Flag', "")
            Nominee_3_Guardian = request.data.get('Nominee_3_Guardian', "")
            Primary_Holder_KYC_Type = request.data.get('Primary_Holder_KYC_Type', "")
            Primary_Holder_CKYC_Number = request.data.get('Primary_Holder_CKYC_Number', "")
            Second_Holder_KYC_Type = request.data.get('Second_Holder_KYC_Type', "")
            Second_Holder_CKYC_Number = request.data.get('Second_Holder_CKYC_Number', "")
            Third_Holder_KYC_Type = request.data.get('Third_Holder_KYC_Type', "")
            Third_Holder_CKYC_Number = request.data.get('Third_Holder_CKYC_Number', "")
            Guardian_KYC_Type = request.data.get('Guardian_KYC_Type', "")
            Guardian_CKYC_Number = request.data.get('Guardian_CKYC_Number', "")
            Primary_Holder_KRA_Exempt_Ref_No = request.data.get('Primary_Holder_KRA_Exempt_Ref_No', "")
            Second_Holder_KRA_Exempt_Ref_No = request.data.get('Second_Holder_KRA_Exempt_Ref_No', "")
            Third_Holder_KRA_Exempt_Ref_No = request.data.get('Third_Holder_KRA_Exempt_Ref_No', "")
            Guardian_Exempt_Ref_No = request.data.get('Guardian_Exempt_Ref_No', "")
            Aadhaar_Updated = request.data.get('Aadhaar_Updated', "")
            Mapin_Id = request.data.get('Mapin_Id', "")
            Paperless_Flag = request.data.get('Paperless_Flag', "")

            LEI_No = request.data.get('LEI_No', "")
            LEI_Validity = request.data.get('LEI_Validity', "")
            Filler_1_Mobile_Declaration_Flag = request.data.get('Filler_1_Mobile_Declaration_Flag', "")
            Filler_2_Email_Declaration_Flag = request.data.get('Filler_2_Email_Declaration_Flag', "")
            Nomination_Opt = request.data.get('Nomination_Opt', "")
            Nomination_Auth_Mode = request.data.get('Nomination_Auth_Mode', "")
            Nominee_PAN1 = request.data.get('Nominee_PAN1', "")
            Nominee_Guardian_PAN1 = request.data.get('Nominee_Guardian_PAN1', "")
            Nominee_PAN2 = request.data.get('Nominee_PAN2', "")
            Nominee_Guardian_PAN2 = request.data.get('Nominee_Guardian_PAN2', "")
            Nominee_PAN3 = request.data.get('Nominee_PAN3', "")
            Nominee_Guardian_PAN3 = request.data.get('Nominee_Guardian_PAN3', "")
            Second_Holder_Email = request.data.get('Second_Holder_Email', "")
            Second_Holder_Email_Declaration = request.data.get('Second_Holder_Email_Declaration', "")
            Second_Holder_Mobile_No = request.data.get('Second_Holder_Mobile_No', "")

            # Field Extraction
            Second_Holder_Mobile_No_Declaration = request.data.get('Second_Holder_Mobile_No_Declaration', "")
            Third_Holder_Email = request.data.get('Third_Holder_Email', "")
            Third_Holder_Email_Declaration = request.data.get('Third_Holder_Email_Declaration', "")
            Third_Holder_Mobile_No = request.data.get('Third_Holder_Mobile_No', "")
            Third_Holder_Mobile_No_Declaration = request.data.get('Third_Holder_Mobile_No_Declaration', "")
            Guardian_Relationship = request.data.get('guardian_relationship', "")
            Filler1 = request.data.get('Filler1', "")  # Reserved for future use
            Filler2 = request.data.get('Filler2', "")  # Reserved for future use
            Filler3 = request.data.get('Filler3', "")  # Reserved for future use
            Trem_conditions = request.data.get('Trem_conditions',"")
            is_submit = request.data.get('is_submit',"")


            # Validation and Conditional Logic
            # if Second_Holder_Mobile_No and not Second_Holder_Mobile_No_Declaration:
            #     return Response({"error": "Second Holder Mobile Declaration is required if Second Holder Mobile No. is provided."}, status=400)

            # if Third_Holder_Email and not Third_Holder_Email_Declaration:
            #     return Response({"error": "Third Holder Email Declaration is required if Third Holder Email is provided."}, status=400)

            # if Third_Holder_Mobile_No and not Third_Holder_Mobile_No_Declaration:
            #     return Response({"error": "Third Holder Mobile Declaration is required if Third Holder Mobile No. is provided."}, status=400)

            # if request.data.get('is_client_minor') == 'Y' and not Guardian_Relationship:
            #     return Response({"error": "Guardian Relationship is required if the client is a minor."}, status=400)

            # Reserved Fields (Filler)
            # You can process Filler1, Filler2, and Filler3 if additional logic is needed in the future.

            agent_id = request.data.get('agent')

            if not agent_id:
                return Response({"detail": "Agent ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                agent = Agent.objects.get(agent_id=agent_id)
            except Agent.DoesNotExist:
                return Response({"detail": "Agent not found."}, status=status.HTTP_404_NOT_FOUND)

            request.data['agent'] = agent.id

            email = data.get("Email")
            
            user_id = request.GET.get("user_id")

            if not user_id:
                if User.objects.filter(username=email).exists():
                    return Response({
                        'status': False,
                        'message': 'A user with this email already exists.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                user_data = {
                    "username": email,
                    "email": email,
                    "password": data.get("password"),
                    "name": data.get("primary_holder_first_name"),
                    "phone": data.get("Indian_Mobile_No"),
                    "role": "U",
                }

                try:
                    user = User.objects.create_user(**user_data)
                    data['user'] = user.id
                except Exception as e:
                    response_data['status'] = False
                    response_data['message'] = f'Error creating user: {str(e)}'
                    return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                clint_code = generate_unique_code()
            else:
                try:
                    user = User.objects.get(id=user_id)
                    user.email = email
                    user.save()
                except User.DoesNotExist:
                    response_data['status'] = False
                    response_data['message'] = "User not found."
                    return Response(response_data, status=status.HTTP_404_NOT_FOUND)
                
            try:
                ucc_obj = Uccregister.objects.get(user=user)

                param_data = ucc_obj.param or {}

                if not param_data.get('clint_code'):
                    clint_code = generate_unique_code()
                    param_data['clint_code'] = clint_code
                    ucc_obj.param = param_data
                    ucc_obj.save()
                else:
                    clint_code = param_data['clint_code']

            except Uccregister.DoesNotExist:
                clint_code = generate_unique_code()

            param_from_request = request.data.get("param", {})
            param_from_request["clint_code"] = clint_code
            request.data["param"] = param_from_request

            serializer = UccregisterSerializer(data=request.data)

            data = request.data
            try:
                data._mutable = True
            except:
                pass
            data['username'] = Email
            data['regn_type'] = regn_type

            if tax_status:
                data['tax_status'] = tax_status
            if occupation_code:
                data['occupation_code'] = occupation_code
            if holding_nature:
                data['holding_code'] = holding_nature
            if Primary_Holder_Exempt_Category:
                data['primary_holder_exempt_category'] = Primary_Holder_Exempt_Category
            if Second_Holder_Exempt_Category:
                data['second_holder_exempt_category'] = Second_Holder_Exempt_Category
            if Third_Holder_Exempt_Category:
                data['third_holder_exempt_category'] = Third_Holder_Exempt_Category
            if Guardian_Exempt_Category:
                data['guardian_exempt_category'] = Guardian_Exempt_Category
            if Client_Type:
                data['client_type'] = Client_Type
            if Account_Type_1:
                data['account_type_1'] = Account_Type_1
            if Account_Type_2:
                data['account_type_2'] = Account_Type_2
            if Account_No_3:
                data['account_type_3'] = Account_No_3
            if Account_Type_4:
                data['account_type_4'] = Account_Type_4
            if Account_Type_5:
                data['account_type_5'] = Account_Type_5
            if Div_Pay_Mode:
                data['div_pay_mode'] = Div_Pay_Mode
            if State:
                data['state_code'] = State
            if Country:
                data['country_code'] = Country
            if Communication_Mode:
                data['communication_mode'] = Communication_Mode
            if Nominee_1_Relationship:
                data['nominee_1_relationship'] = Nominee_1_Relationship
            if Nominee_2_Relationship:
                data['nominee_2_relationship'] = Nominee_2_Relationship
            if Nominee_3_Relationship:
                data['nominee_3_relationship'] = Nominee_3_Relationship
            if Guardian_Relationship:
                data['guardian_relationship'] = Guardian_Relationship
            if Primary_Holder_KYC_Type:
                data['primary_holder_kyc_type'] = Primary_Holder_KYC_Type
            if Second_Holder_KYC_Type:
                data['second_holder_kyc_type'] = Second_Holder_KYC_Type
            if Third_Holder_KYC_Type:
                data['third_holder_kyc_type'] = Third_Holder_KYC_Type
            if Guardian_KYC_Type:
                data['guardian_holder_kyc_type'] = Guardian_KYC_Type
            if Paperless_Flag:
                data['paperless_flag'] = Paperless_Flag
            if Filler_1_Mobile_Declaration_Flag:
                data['filler_1_mobile_declaration_flag'] = Filler_1_Mobile_Declaration_Flag
            if Filler_2_Email_Declaration_Flag:
                data['filler_2_email_declaration_flag'] = Filler_2_Email_Declaration_Flag
            if Second_Holder_Email_Declaration:
                data['second_holder_email_declaration'] = Second_Holder_Email_Declaration
            if Second_Holder_Mobile_No_Declaration:
                data['second_holder_mobile_declaration'] = Second_Holder_Mobile_No_Declaration
            if Third_Holder_Mobile_No_Declaration:
                data['third_holder_mobile_declaration'] = Third_Holder_Mobile_No_Declaration
            if Third_Holder_Email_Declaration:
                data['third_holder_email_declaration'] = Third_Holder_Email_Declaration
            if Nomination_Auth_Mode:
                data['nomination_auth_mode'] = Nomination_Auth_Mode
            if Indian_Mobile_No:
                data['mobile'] = Indian_Mobile_No
            if Email:
                data['email'] = Email
            if Trem_conditions:
                data['Trem_conditions'] = Trem_conditions
            if is_submit:
                data['is_submit'] = is_submit

            if 'is_submit' not in data:
                data['is_submit'] = False
            
            id = request.GET.get('id')

            if is_submit == "true":

                if id:
                    try:
                        user = Uccregister.objects.get(id=id)
                    except Uccregister.DoesNotExist:
                        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
                    if user is not None:
                                primary_holder_first_name = user.param['primary_holder_first_name']
                                primary_holder_middle_name = user.param['primary_holder_middle_name']
                                primary_holder_last_name = user.param['primary_holder_last_name']
                                tax_status = user.param['tax_status']
                                gender = user.param['gender']
                                primary_holder_dob_incorporation = user.param['primary_holder_dob_incorporation']
                                occupation_code = user.param['occupation_code']
                                holding_nature = user.param['holding_nature']
                                Second_Holder_First_Name = user.param['Second_Holder_First_Name']
                                Second_Holder_Middle_Name = user.param['Second_Holder_Middle_Name']
                                Second_Holder_Last_Name = user.param['Second_Holder_Last_Name']
                                Third_Holder_First_Name = user.param['Third_Holder_First_Name']
                                Third_Holder_Middle_Name = user.param['Third_Holder_Middle_Name']
                                Third_Holder_Last_Name = user.param['Third_Holder_Last_Name']
                                Second_Holder_DOB = user.param['Second_Holder_DOB']
                                Third_Holder_DOB = user.param['Third_Holder_DOB']
                                Guardian_First_Name = user.param['Guardian_First_Name']
                                Guardian_Middle_Name = user.param['Guardian_Middle_Name']
                                Guardian_Last_Name = user.param['Guardian_Last_Name']
                                Guardian_DOB = user.param['Guardian_DOB']
                                Primary_Holder_PAN_Exempt = user.param['Primary_Holder_PAN_Exempt']
                                Second_Holder_PAN_Exempt = user.param['Second_Holder_PAN_Exempt']
                                Third_Holder_PAN_Exempt = user.param['Third_Holder_PAN_Exempt']
                                Guardian_PAN_Exempt = user.param['Guardian_PAN_Exempt']
                                Primary_Holder_PAN = user.param['Primary_Holder_PAN']
                                Second_Holder_PAN = user.param['Second_Holder_PAN']
                                Third_Holder_PAN = user.param['Third_Holder_PAN']
                                Guardian_PAN = user.param['Guardian_PAN']
                                Primary_Holder_Exempt_Category = user.param['Primary_Holder_Exempt_Category']
                                Second_Holder_Exempt_Category = user.param['Second_Holder_Exempt_Category']
                                Third_Holder_Exempt_Category = user.param['Third_Holder_Exempt_Category']
                                Guardian_Exempt_Category = user.param['Guardian_Exempt_Category']
                                Client_Type = user.param['Client_Type']
                                pms = user.param['pms']
                                Default_DP = user.param['Default_DP']
                                cdsl_dpid = user.param['cdsl_dpid']
                                cdslctdid = user.param['cdslctdid']
                                CMBP_Id = user.param['CMBP_Id']
                                nsdl_dpid = user.param['nsdl_dpid']
                                nsdlcltid = user.param['nsdlcltid']
                                Account_Type_1 = user.param['Account_Type_1']
                                Account_No_1 = user.param['Account_No_1']
                                MICR_No_1 = user.param['MICR_No_1']
                                IFSC_Code_1 = user.param['IFSC_Code_1']
                                Default_Bank_Flag_1 = user.param['Default_Bank_Flag_1']
                                Account_Type_2 = user.param['Account_Type_2']
                                Account_No_2 = user.param['Account_No_2']
                                MICR_No_2 = user.param['MICR_No_2']
                                IFSC_Code_2 = user.param['IFSC_Code_2']
                                Default_Bank_Flag_2 = user.param['Default_Bank_Flag_2']
                                Account_Type_3 = user.param['Account_Type_3']
                                Account_No_3 = user.param['Account_No_3']
                                MICR_No_3 = user.param['MICR_No_3']
                                IFSC_Code_3 = user.param['IFSC_Code_3']
                                Default_Bank_Flag_3 = user.param['Default_Bank_Flag_3']
                                Account_Type_4 = user.param['Account_Type_4']
                                Account_No_4 = user.param['Account_No_4']
                                MICR_No_4 = user.param['MICR_No_4']
                                IFSC_Code_4 = user.param['IFSC_Code_4']
                                Default_Bank_Flag_4 = user.param['Default_Bank_Flag_4']
                                Account_Type_5 = user.param['Account_Type_5']
                                Account_No_5 = user.param['Account_No_5']
                                MICR_No_5 = user.param['MICR_No_5']
                                IFSC_Code_5 = user.param['IFSC_Code_5']
                                Default_Bank_Flag_5 = user.param['Default_Bank_Flag_5']
                                Cheque_Name = user.param['Cheque_Name']
                                Div_Pay_Mode = user.param['Div_Pay_Mode']
                                Address_1 = user.param['Address_1']
                                Address_2 = user.param['Address_2']
                                Address_3 = user.param['Address_3']
                                City = user.param['City']
                                State = user.param['State']
                                Pincode = user.param['Pincode']
                                Country = user.param['Country']
                                Resi_Phone = user.param['Resi_Phone']
                                Resi_Fax = user.param['Resi_Fax']
                                Office_Phone = user.param['Office_Phone']
                                Office_Fax = user.param['Office_Fax']
                                Email = user.param['Email']
                                Communication_Mode = user.param['Communication_Mode']
                                Foreign_Address_1 = user.param['Foreign_Address_1']
                                Foreign_Address_2 = user.param['Foreign_Address_2']
                                Foreign_Address_3 = user.param['Foreign_Address_3']
                                Foreign_Address_City = user.param['Foreign_Address_City']
                                Foreign_Address_Pincode = user.param['Foreign_Address_Pincode']
                                Foreign_Address_State = user.param['Foreign_Address_State']
                                Foreign_Address_Country = user.param['Foreign_Address_Country']
                                Foreign_Address_Resi_Phone = user.param['Foreign_Address_Resi_Phone']
                                Foreign_Address_Fax = user.param['Foreign_Address_Fax']
                                Foreign_Address_Off_Phone = user.param['Foreign_Address_Off_Phone']
                                Foreign_Address_Off_Fax = user.param['Foreign_Address_Off_Fax']
                                Indian_Mobile_No = user.param['Indian_Mobile_No']
                                Nominee_1_Name = user.param['Nominee_1_Name']
                                Nominee_1_Relationship = user.param['Nominee_1_Relationship']
                                Nominee_1_Applicable = user.param['Nominee_1_Applicable']
                                Nominee_1_Minor_Flag = user.param['Nominee_1_Minor_Flag']
                                Nominee_1_DOB = user.param['Nominee_1_DOB']
                                Nominee_1_Guardian = user.param['Nominee_1_Guardian']
                                Nominee_2_Name = user.param['Nominee_2_Name']
                                Nominee_2_Relationship = user.param['Nominee_2_Relationship']
                                Nominee_2_Applicable = user.param['Nominee_2_Applicable']
                                Nominee_2_DOB = user.param['Nominee_2_DOB']
                                Nominee_2_Minor_Flag = user.param['Nominee_2_Minor_Flag']
                                Nominee_2_Guardian = user.param['Nominee_2_Guardian']
                                Nominee_3_Name = user.param['Nominee_3_Name']
                                Nominee_3_Relationship = user.param['Nominee_3_Relationship']
                                Nominee_3_Applicable = user.param['Nominee_3_Applicable']
                                Nominee_3_DOB = user.param['Nominee_3_DOB']
                                Nominee_3_Minor_Flag = user.param['Nominee_3_Minor_Flag']
                                Nominee_3_Guardian = user.param['Nominee_3_Guardian']
                                Primary_Holder_KYC_Type = user.param['Primary_Holder_KYC_Type']
                                Primary_Holder_CKYC_Number = user.param['Primary_Holder_CKYC_Number']
                                Second_Holder_KYC_Type = user.param['Second_Holder_KYC_Type']
                                Second_Holder_CKYC_Number = user.param['Second_Holder_CKYC_Number']
                                Third_Holder_KYC_Type = user.param['Third_Holder_KYC_Type']
                                Third_Holder_CKYC_Number = user.param['Third_Holder_CKYC_Number']
                                Guardian_KYC_Type = user.param['Guardian_KYC_Type']
                                Guardian_CKYC_Number = user.param['Guardian_CKYC_Number']
                                Primary_Holder_KRA_Exempt_Ref_No = user.param['Primary_Holder_KRA_Exempt_Ref_No']
                                Second_Holder_KRA_Exempt_Ref_No = user.param['Second_Holder_KRA_Exempt_Ref_No']
                                Third_Holder_KRA_Exempt_Ref_No = user.param['Third_Holder_KRA_Exempt_Ref_No']
                                Guardian_Exempt_Ref_No = user.param['Guardian_Exempt_Ref_No']
                                Aadhaar_Updated = user.param['Aadhaar_Updated']
                                Mapin_Id = user.param['Mapin_Id']
                                Paperless_Flag = user.param['Paperless_Flag']
                                LEI_No = user.param['LEI_No']
                                LEI_Validity = user.param['LEI_Validity']
                                Filler_1_Mobile_Declaration_Flag=user.param['Filler_1_Mobile_Declaration_Flag']
                                Filler_2_Email_Declaration_Flag=user.param['Filler_2_Email_Declaration_Flag']
                                Nomination_Opt=user.param['Nomination_Opt']
                                Nomination_Auth_Mode=user.param['Nomination_Auth_Mode']
                                Filler1 = Filler1
                                Filler2 = Filler2
                                Filler3 = Filler3

                print("ifffffffffffffffffffffffffffffffffffffffffffffffffffff")
                conn = http.client.HTTPSConnection(base_url)
                payload = json.dumps({
                "UserId": loginid,
                "MemberCode": membercode,
                "Password": password,
                "RegnType": regn_type,
                # "Param": "PH109999|Gaurav||Verma|01|F|24/01/1991|02|SI|||||||||||||N||||ANSPG9326N||||||||P||||||||SB|30191531880|382002101|SBIN0000362|Y|||||||||||||||||||||ASHOKKUMAR PRAVINBHAI GADANI|02|3976|SUTARIYA STREET|TALAV KANTHEROJKA|AHMEDABAD|GU|382460|India|9999999999||||abcdef@abc.com|E||||||||||||9999999990|||||||||||||||||||K||||||||||||||Z|||SE|SE|N|O||||||||||||||",
                # "Param" : f"{clint_code}|{primary_holder_first_name}|{primary_holder_middle_name}|{primary_holder_last_name}|{tax_status}|{gender}|{primary_holder_dob_incorporation}|{occupation_code}|{holding_nature}|{Second_Holder_First_Name}|{Second_Holder_Middle_Name}|{Second_Holder_Last_Name}|{Second_Holder_DOB}|{Third_Holder_First_Name}|{Third_Holder_Middle_Name}|{Third_Holder_Last_Name}|{Third_Holder_DOB}|{Guardian_First_Name}|{Guardian_Middle_Name}|{Guardian_Last_Name}|{Guardian_DOB}|{Primary_Holder_PAN_Exempt}|{Second_Holder_PAN_Exempt}|{Third_Holder_PAN_Exempt}|{Guardian_PAN_Exempt}|{Primary_Holder_PAN}|{Second_Holder_PAN}|{Third_Holder_PAN}|{Guardian_PAN}|{Primary_Holder_Exempt_Category}|{Second_Holder_Exempt_Category}|{Third_Holder_Exempt_Category}|{Guardian_Exempt_Category}|{Client_Type}|{pms}|{Default_DP}|{cdsl_dpid}|{cdslctdid}|{CMBP_Id}|{nsdl_dpid}|{nsdlcltid}|{Account_Type_1}|{Account_No_1}|{MICR_No_1}|{IFSC_Code_1}|{Default_Bank_Flag_1}|{Account_Type_2}|{Account_No_2}|{MICR_No_2}|{IFSC_Code_2}|{Default_Bank_Flag_2}|{Account_Type_3}|{Account_No_3}|{MICR_No_3}|{IFSC_Code_3}|{Default_Bank_Flag_3}|{Account_Type_4}|{Account_No_4}|{MICR_No_4}|{IFSC_Code_4}|{Default_Bank_Flag_4}|{Account_Type_5}|{Account_No_5}|{MICR_No_5}|{IFSC_Code_5}|{Default_Bank_Flag_5}|{Cheque_Name}|{Div_Pay_Mode}|{Address_1}|{Address_2}|{Address_3}|{City}|{State}|{Pincode}|{Country}|{Resi_Phone}|{Resi_Fax}|{Office_Phone}|{Office_Fax}|{Email}|{Communication_Mode}|{Foreign_Address_1}|{Foreign_Address_2}|{Foreign_Address_3}|{Foreign_Address_City}|{Foreign_Address_Pincode}|{Foreign_Address_State}|{Foreign_Address_Country}|{Foreign_Address_Resi_Phone}|{Foreign_Address_Fax}|{Foreign_Address_Off_Phone}|{Foreign_Address_Off_Fax}|{Indian_Mobile_No}|{Nominee_1_Name}|{Nominee_1_Relationship}|{Nominee_1_Applicable}|{Nominee_1_Minor_Flag}|{Nominee_1_DOB}|{Nominee_1_Guardian}|{Nominee_2_Name}|{Nominee_2_Relationship}|{Nominee_2_Applicable}|{Nominee_2_DOB}|{Nominee_2_Minor_Flag}|{Nominee_2_Guardian}|{Nominee_3_Name}|{Nominee_3_Relationship}|{Nominee_3_Applicable}|{Nominee_3_DOB}|{Nominee_3_Minor_Flag}|{Nominee_3_Guardian}|{Primary_Holder_KYC_Type}|{Primary_Holder_CKYC_Number}|{Second_Holder_KYC_Type}|{Second_Holder_CKYC_Number}|{Third_Holder_KYC_Type}|{Third_Holder_CKYC_Number}|{Guardian_KYC_Type}|{Guardian_CKYC_Number}|{Primary_Holder_KRA_Exempt_Ref_No}|{Second_Holder_KRA_Exempt_Ref_No}|{Third_Holder_KRA_Exempt_Ref_No}|{Guardian_Exempt_Ref_No}|{Aadhaar_Updated}|{Mapin_Id}|{Paperless_Flag}|{LEI_No}|{LEI_Validity}|{Filler_1_Mobile_Declaration_Flag}|{Filler_2_Email_Declaration_Flag}|{Nomination_Opt}|{Nomination_Auth_Mode}|{Nominee_PAN1}|{Nominee_Guardian_PAN1}|{Nominee_PAN2}|{Nominee_Guardian_PAN2}|{Nominee_PAN3}|{Nominee_Guardian_PAN3}|{Second_Holder_Email}|{Second_Holder_Email_Declaration}|{Second_Holder_Mobile_No}|{Second_Holder_Mobile_No_Declaration}|{Third_Holder_Email}|{Third_Holder_Email_Declaration}|{Third_Holder_Mobile_No}|{Third_Holder_Mobile_No_Declaration}|{Guardian_Relationship}",

                # "Param": "PH109933|Gaurav|||01|M|01/01/1990|01|SI|||||||||||||N||||ABCPD1234F||||||||P||||||||SB|11415123||HDFC0000002|Y|||||||||||||||||||||FirstNameLastName|01|ADD1|ADD2|ADD3|MUMBAI|MA|400001|INDIA|||||test@test.com|P||||||||||||9982908095|NomineeName1|01|100|N|||||||||||||||K||||||||||||N||P|||SE|SE|Y|E||||||||||||||||||",
                "Param": f"{clint_code}|{primary_holder_first_name}|{primary_holder_middle_name}|{primary_holder_last_name}|{tax_status}|{gender}|{primary_holder_dob_incorporation}|{occupation_code}|{holding_nature}|{Second_Holder_First_Name}|{Second_Holder_Middle_Name}|{Second_Holder_Last_Name}|{Third_Holder_First_Name}|{Third_Holder_Middle_Name}|{Third_Holder_Last_Name}|{Second_Holder_DOB}|{Third_Holder_DOB}|{Guardian_First_Name}|{Guardian_Middle_Name}|{Guardian_Last_Name}|{Guardian_DOB}|{Primary_Holder_PAN_Exempt}|{Second_Holder_PAN_Exempt}|{Third_Holder_PAN_Exempt}|{Guardian_PAN_Exempt}|{Primary_Holder_PAN}|{Second_Holder_PAN}|{Third_Holder_PAN}|{Guardian_PAN}|{Primary_Holder_Exempt_Category}|{Second_Holder_Exempt_Category}|{Third_Holder_Exempt_Category}|{Guardian_Exempt_Category}|{Client_Type}|{pms}|{Default_DP}|{cdsl_dpid}|{cdslctdid}|{CMBP_Id}|{nsdl_dpid}|{nsdlcltid}|{Account_Type_1}|{Account_No_1}|{MICR_No_1}|{IFSC_Code_1}|{Default_Bank_Flag_1}|{Account_Type_2}|{Account_No_2}|{MICR_No_2}|{IFSC_Code_2}|{Default_Bank_Flag_2}|{Account_Type_3}|{Account_No_3}|{MICR_No_3}|{IFSC_Code_3}|{Default_Bank_Flag_3}|{Account_Type_4}|{Account_No_4}|{MICR_No_4}|{IFSC_Code_4}|{Default_Bank_Flag_4}|{Account_Type_5}|{Account_No_5}|{MICR_No_5}|{IFSC_Code_5}|{Default_Bank_Flag_5}|{Cheque_Name}|{Div_Pay_Mode}|{Address_1}|{Address_2}|{Address_3}|{City}|{State}|{Pincode}|{Country}|{Resi_Phone}|{Resi_Fax}|{Office_Phone}|{Office_Fax}|{Email}|{Communication_Mode}|{Foreign_Address_1}|{Foreign_Address_2}|{Foreign_Address_3}|{Foreign_Address_City}|{Foreign_Address_Pincode}|{Foreign_Address_State}|{Foreign_Address_Country}|{Foreign_Address_Resi_Phone}|{Foreign_Address_Fax}|{Foreign_Address_Off_Phone}|{Foreign_Address_Off_Fax}|{Indian_Mobile_No}|{Nominee_1_Name}|{Nominee_1_Relationship}|{Nominee_1_Applicable}|{Nominee_1_Minor_Flag}|{Nominee_1_DOB}|{Nominee_1_Guardian}|{Nominee_2_Name}|{Nominee_2_Relationship}|{Nominee_2_Applicable}|{Nominee_2_DOB}|{Nominee_2_Minor_Flag}|{Nominee_2_Guardian}|{Nominee_3_Name}|{Nominee_3_Relationship}|{Nominee_3_Applicable}|{Nominee_3_DOB}|{Nominee_3_Minor_Flag}|{Nominee_3_Guardian}|{Primary_Holder_KYC_Type}|{Primary_Holder_CKYC_Number}|{Second_Holder_KYC_Type}|{Second_Holder_CKYC_Number}|{Third_Holder_KYC_Type}|{Third_Holder_CKYC_Number}|{Guardian_KYC_Type}|{Guardian_CKYC_Number}|{Primary_Holder_KRA_Exempt_Ref_No}|{Second_Holder_KRA_Exempt_Ref_No}|{Third_Holder_KRA_Exempt_Ref_No}|{Guardian_Exempt_Ref_No}|{Aadhaar_Updated}|{Mapin_Id}|{Paperless_Flag}|{LEI_No}|{LEI_Validity}|{Filler_1_Mobile_Declaration_Flag}|{Filler_2_Email_Declaration_Flag}|{Nomination_Opt}|{Nomination_Auth_Mode}|{Nominee_PAN1}|{Nominee_Guardian_PAN1}|{Nominee_PAN2}|{Nominee_Guardian_PAN2}|{Nominee_PAN3}|{Nominee_Guardian_PAN3}|{Second_Holder_Email}|{Second_Holder_Email_Declaration}|{Second_Holder_Mobile_No}|{Second_Holder_Mobile_No_Declaration}|{Third_Holder_Email}|{Third_Holder_Email_Declaration}|{Third_Holder_Mobile_No}|{Third_Holder_Mobile_No_Declaration}|{Guardian_Relationship}|{Filler1}|{Filler2}|{Filler3}",
                "Filler1": "",
                "Filler2": ""
                })   
                print(payload, 'CCCCCCCCCCCCCCC') 
                headers = {
                'Content-Type': 'application/json'
                }
                conn.request("POST", "/BSEMFWEBAPI/UCCAPI/UCCRegistration", payload, headers)
                res = conn.getresponse()
                api_data = res.read().decode()
                print(res.status, res.reason, api_data, 'NNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNN')
                data_dict = json.loads(api_data)
                if data_dict['Status'] == '0':
                    data['param'] = {
                                    "clint_code": clint_code,
                                    "primary_holder_first_name": primary_holder_first_name,
                                    "primary_holder_middle_name": primary_holder_middle_name,
                                    "primary_holder_last_name": primary_holder_last_name,
                                    "tax_status": tax_status,
                                    "gender": gender,
                                    "primary_holder_dob_incorporation": primary_holder_dob_incorporation,
                                    "occupation_code": occupation_code,
                                    "holding_nature": holding_nature,
                                    "Second_Holder_First_Name": Second_Holder_First_Name,
                                    "Second_Holder_Middle_Name": Second_Holder_Middle_Name,
                                    "Second_Holder_Last_Name": Second_Holder_Last_Name,
                                    "Third_Holder_First_Name": Third_Holder_First_Name,
                                    "Third_Holder_Middle_Name": Third_Holder_Middle_Name,
                                    "Third_Holder_Last_Name": Third_Holder_Last_Name,
                                    "Second_Holder_DOB": Second_Holder_DOB,
                                    "Third_Holder_DOB": Third_Holder_DOB,
                                    "Guardian_First_Name": Guardian_First_Name,
                                    "Guardian_Middle_Name": Guardian_Middle_Name,
                                    "Guardian_Last_Name": Guardian_Last_Name,
                                    "Guardian_DOB": Guardian_DOB,
                                    "Primary_Holder_PAN_Exempt": Primary_Holder_PAN_Exempt,
                                    "Second_Holder_PAN_Exempt": Second_Holder_PAN_Exempt,
                                    "Third_Holder_PAN_Exempt": Third_Holder_PAN_Exempt,
                                    "Guardian_PAN_Exempt": Guardian_PAN_Exempt,
                                    "Primary_Holder_PAN": Primary_Holder_PAN,
                                    "Second_Holder_PAN": Second_Holder_PAN,
                                    "Third_Holder_PAN": Third_Holder_PAN,
                                    "Guardian_PAN": Guardian_PAN,
                                    "Primary_Holder_Exempt_Category": Primary_Holder_Exempt_Category,
                                    "Second_Holder_Exempt_Category": Second_Holder_Exempt_Category,
                                    "Third_Holder_Exempt_Category": Third_Holder_Exempt_Category,
                                    "Guardian_Exempt_Category": Guardian_Exempt_Category,
                                    "Client_Type": Client_Type,
                                    "pms": pms,
                                    "Default_DP": Default_DP,
                                    "cdsl_dpid": cdsl_dpid,
                                    "cdslctdid": cdslctdid,
                                    "CMBP_Id": CMBP_Id,
                                    "nsdl_dpid": nsdl_dpid,
                                    "nsdlcltid": nsdlcltid,
                                    "Account_Type_1": Account_Type_1,
                                    "Account_No_1": Account_No_1,
                                    "MICR_No_1": MICR_No_1,
                                    "IFSC_Code_1": IFSC_Code_1,
                                    "Default_Bank_Flag_1": Default_Bank_Flag_1,
                                    "Account_Type_2": Account_Type_2,
                                    "Account_No_2": Account_No_2,
                                    "MICR_No_2": MICR_No_2,
                                    "IFSC_Code_2": IFSC_Code_2,
                                    "Default_Bank_Flag_2": Default_Bank_Flag_2,
                                    "Account_Type_3": Account_Type_3,
                                    "Account_No_3": Account_No_3,
                                    "MICR_No_3": MICR_No_3,
                                    "IFSC_Code_3": IFSC_Code_3,
                                    "Default_Bank_Flag_3": Default_Bank_Flag_3,
                                    "Account_Type_4": Account_Type_4,
                                    "Account_No_4": Account_No_4,
                                    "MICR_No_4": MICR_No_4,
                                    "IFSC_Code_4": IFSC_Code_4,
                                    "Default_Bank_Flag_4": Default_Bank_Flag_4,
                                    "Account_Type_5": Account_Type_5,
                                    "Account_No_5": Account_No_5,
                                    "MICR_No_5": MICR_No_5,
                                    "IFSC_Code_5": IFSC_Code_5,
                                    "Default_Bank_Flag_5": Default_Bank_Flag_5,
                                    "Cheque_Name": Cheque_Name,
                                    "Div_Pay_Mode": Div_Pay_Mode,
                                    "Address_1": Address_1,
                                    "Address_2": Address_2,
                                    "Address_3": Address_3,
                                    "City": City,
                                    "State": State,
                                    "Pincode": Pincode,
                                    "Country": Country,
                                    "Resi_Phone": Resi_Phone,
                                    "Resi_Fax": Resi_Fax,
                                    "Office_Phone": Office_Phone,
                                    "Office_Fax": Office_Fax,
                                    "Email": Email,
                                    "Communication_Mode": Communication_Mode,
                                    "Foreign_Address_1": Foreign_Address_1,
                                    "Foreign_Address_2": Foreign_Address_2,
                                    "Foreign_Address_3": Foreign_Address_3,
                                    "Foreign_Address_City": Foreign_Address_City,
                                    "Foreign_Address_Pincode": Foreign_Address_Pincode,
                                    "Foreign_Address_State": Foreign_Address_State,
                                    "Foreign_Address_Country": Foreign_Address_Country,
                                    "Foreign_Address_Resi_Phone": Foreign_Address_Resi_Phone,
                                    "Foreign_Address_Fax": Foreign_Address_Fax,
                                    "Foreign_Address_Off_Phone": Foreign_Address_Off_Phone,
                                    "Foreign_Address_Off_Fax": Foreign_Address_Off_Fax,
                                    "Indian_Mobile_No": Indian_Mobile_No,
                                    "Nominee_1_Name": Nominee_1_Name,
                                    "Nominee_1_Relationship": Nominee_1_Relationship,
                                    "Nominee_1_Applicable": Nominee_1_Applicable,
                                    "Nominee_1_Minor_Flag": Nominee_1_Minor_Flag,
                                    "Nominee_1_DOB": Nominee_1_DOB,
                                    "Nominee_1_Guardian": Nominee_1_Guardian,
                                    "Nominee_2_Name": Nominee_2_Name,
                                    "Nominee_2_Relationship": Nominee_2_Relationship,
                                    "Nominee_2_Applicable": Nominee_2_Applicable,
                                    "Nominee_2_DOB": Nominee_2_DOB,
                                    "Nominee_2_Minor_Flag": Nominee_2_Minor_Flag,
                                    "Nominee_2_Guardian": Nominee_2_Guardian,
                                    "Nominee_3_Name": Nominee_3_Name,
                                    "Nominee_3_Relationship": Nominee_3_Relationship,
                                    "Nominee_3_Applicable": Nominee_3_Applicable,
                                    "Nominee_3_DOB": Nominee_3_DOB,
                                    "Nominee_3_Minor_Flag": Nominee_3_Minor_Flag,
                                    "Nominee_3_Guardian": Nominee_3_Guardian,
                                    "Primary_Holder_KYC_Type": Primary_Holder_KYC_Type,
                                    "Primary_Holder_CKYC_Number": Primary_Holder_CKYC_Number,
                                    "Second_Holder_KYC_Type": Second_Holder_KYC_Type,
                                    "Second_Holder_CKYC_Number": Second_Holder_CKYC_Number,
                                    "Third_Holder_KYC_Type": Third_Holder_KYC_Type,
                                    "Third_Holder_CKYC_Number": Third_Holder_CKYC_Number,
                                    "Guardian_KYC_Type": Guardian_KYC_Type,
                                    "Guardian_CKYC_Number": Guardian_CKYC_Number,
                                    "Primary_Holder_KRA_Exempt_Ref_No": Primary_Holder_KRA_Exempt_Ref_No,
                                    "Second_Holder_KRA_Exempt_Ref_No": Second_Holder_KRA_Exempt_Ref_No,
                                    "Third_Holder_KRA_Exempt_Ref_No": Third_Holder_KRA_Exempt_Ref_No,
                                    "Guardian_Exempt_Ref_No": Guardian_Exempt_Ref_No,
                                    "Aadhaar_Updated": Aadhaar_Updated,
                                    "Mapin_Id": Mapin_Id,
                                    "Paperless_Flag": Paperless_Flag,
                                    "LEI_No": LEI_No,
                                    "LEI_Validity": LEI_Validity,
                                    "Filler_1_Mobile_Declaration_Flag": Filler_1_Mobile_Declaration_Flag,
                                    "Filler_2_Email_Declaration_Flag": Filler_2_Email_Declaration_Flag,
                                    "Nomination_Opt": Nomination_Opt,
                                    "Nomination_Auth_Mode": Nomination_Auth_Mode,
                                    "Nominee_PAN1": Nominee_PAN1,
                                    "Nominee_Guardian_PAN1": Nominee_Guardian_PAN1,
                                    "Nominee_PAN2": Nominee_PAN2,
                                    "Nominee_Guardian_PAN2": Nominee_Guardian_PAN2,
                                    "Nominee_PAN3": Nominee_PAN3,
                                    "Nominee_Guardian_PAN3": Nominee_Guardian_PAN3,
                                    "Second_Holder_Email": Second_Holder_Email,
                                    "Second_Holder_Email_Declaration": Second_Holder_Email_Declaration,
                                    "Second_Holder_Mobile_No": Second_Holder_Mobile_No,
                                    "Second_Holder_Mobile_No_Declaration": Second_Holder_Mobile_No_Declaration,
                                    "Third_Holder_Email": Third_Holder_Email,
                                    "Third_Holder_Email_Declaration": Third_Holder_Email_Declaration,
                                    "Third_Holder_Mobile_No": Third_Holder_Mobile_No,
                                    "Third_Holder_Mobile_No_Declaration": Third_Holder_Mobile_No_Declaration,
                                    "guardian_relationship": Guardian_Relationship,
                                }
                else:
                    response_data['status'] = False
                    response_data['message'] = 'Error creating user'
                    response_data['errors'] = data_dict['Remarks']
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
                
            if is_submit == "false":
                print("elseeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee")
                data['param'] = {
                                    "primary_holder_first_name": primary_holder_first_name,
                                    "primary_holder_middle_name": primary_holder_middle_name,
                                    "primary_holder_last_name": primary_holder_last_name,
                                    "tax_status": tax_status,
                                    "gender": gender,
                                    "primary_holder_dob_incorporation": primary_holder_dob_incorporation,
                                    "occupation_code": occupation_code,
                                    "holding_nature": holding_nature,
                                    "Second_Holder_First_Name": Second_Holder_First_Name,
                                    "Second_Holder_Middle_Name": Second_Holder_Middle_Name,
                                    "Second_Holder_Last_Name": Second_Holder_Last_Name,
                                    "Third_Holder_First_Name": Third_Holder_First_Name,
                                    "Third_Holder_Middle_Name": Third_Holder_Middle_Name,
                                    "Third_Holder_Last_Name": Third_Holder_Last_Name,
                                    "Second_Holder_DOB": Second_Holder_DOB,
                                    "Third_Holder_DOB": Third_Holder_DOB,
                                    "Guardian_First_Name": Guardian_First_Name,
                                    "Guardian_Middle_Name": Guardian_Middle_Name,
                                    "Guardian_Last_Name": Guardian_Last_Name,
                                    "Guardian_DOB": Guardian_DOB,
                                    "Primary_Holder_PAN_Exempt": Primary_Holder_PAN_Exempt,
                                    "Second_Holder_PAN_Exempt": Second_Holder_PAN_Exempt,
                                    "Third_Holder_PAN_Exempt": Third_Holder_PAN_Exempt,
                                    "Guardian_PAN_Exempt": Guardian_PAN_Exempt,
                                    "Primary_Holder_PAN": Primary_Holder_PAN,
                                    "Second_Holder_PAN": Second_Holder_PAN,
                                    "Third_Holder_PAN": Third_Holder_PAN,
                                    "Guardian_PAN": Guardian_PAN,
                                    "Primary_Holder_Exempt_Category": Primary_Holder_Exempt_Category,
                                    "Second_Holder_Exempt_Category": Second_Holder_Exempt_Category,
                                    "Third_Holder_Exempt_Category": Third_Holder_Exempt_Category,
                                    "Guardian_Exempt_Category": Guardian_Exempt_Category,
                                    "Client_Type": Client_Type,
                                    "pms": pms,
                                    "Default_DP": Default_DP,
                                    "cdsl_dpid": cdsl_dpid,
                                    "cdslctdid": cdslctdid,
                                    "CMBP_Id": CMBP_Id,
                                    "nsdl_dpid": nsdl_dpid,
                                    "nsdlcltid": nsdlcltid,
                                    "Account_Type_1": Account_Type_1,
                                    "Account_No_1": Account_No_1,
                                    "MICR_No_1": MICR_No_1,
                                    "IFSC_Code_1": IFSC_Code_1,
                                    "Default_Bank_Flag_1": Default_Bank_Flag_1,
                                    "Account_Type_2": Account_Type_2,
                                    "Account_No_2": Account_No_2,
                                    "MICR_No_2": MICR_No_2,
                                    "IFSC_Code_2": IFSC_Code_2,
                                    "Default_Bank_Flag_2": Default_Bank_Flag_2,
                                    "Account_Type_3": Account_Type_3,
                                    "Account_No_3": Account_No_3,
                                    "MICR_No_3": MICR_No_3,
                                    "IFSC_Code_3": IFSC_Code_3,
                                    "Default_Bank_Flag_3": Default_Bank_Flag_3,
                                    "Account_Type_4": Account_Type_4,
                                    "Account_No_4": Account_No_4,
                                    "MICR_No_4": MICR_No_4,
                                    "IFSC_Code_4": IFSC_Code_4,
                                    "Default_Bank_Flag_4": Default_Bank_Flag_4,
                                    "Account_Type_5": Account_Type_5,
                                    "Account_No_5": Account_No_5,
                                    "MICR_No_5": MICR_No_5,
                                    "IFSC_Code_5": IFSC_Code_5,
                                    "Default_Bank_Flag_5": Default_Bank_Flag_5,
                                    "Cheque_Name": Cheque_Name,
                                    "Div_Pay_Mode": Div_Pay_Mode,
                                    "Address_1": Address_1,
                                    "Address_2": Address_2,
                                    "Address_3": Address_3,
                                    "City": City,
                                    "State": State,
                                    "Pincode": Pincode,
                                    "Country": Country,
                                    "Resi_Phone": Resi_Phone,
                                    "Resi_Fax": Resi_Fax,
                                    "Office_Phone": Office_Phone,
                                    "Office_Fax": Office_Fax,
                                    "Email": Email,
                                    "Communication_Mode": Communication_Mode,
                                    "Foreign_Address_1": Foreign_Address_1,
                                    "Foreign_Address_2": Foreign_Address_2,
                                    "Foreign_Address_3": Foreign_Address_3,
                                    "Foreign_Address_City": Foreign_Address_City,
                                    "Foreign_Address_Pincode": Foreign_Address_Pincode,
                                    "Foreign_Address_State": Foreign_Address_State,
                                    "Foreign_Address_Country": Foreign_Address_Country,
                                    "Foreign_Address_Resi_Phone": Foreign_Address_Resi_Phone,
                                    "Foreign_Address_Fax": Foreign_Address_Fax,
                                    "Foreign_Address_Off_Phone": Foreign_Address_Off_Phone,
                                    "Foreign_Address_Off_Fax": Foreign_Address_Off_Fax,
                                    "Indian_Mobile_No": Indian_Mobile_No,
                                    "Nominee_1_Name": Nominee_1_Name,
                                    "Nominee_1_Relationship": Nominee_1_Relationship,
                                    "Nominee_1_Applicable": Nominee_1_Applicable,
                                    "Nominee_1_Minor_Flag": Nominee_1_Minor_Flag,
                                    "Nominee_1_DOB": Nominee_1_DOB,
                                    "Nominee_1_Guardian": Nominee_1_Guardian,
                                    "Nominee_2_Name": Nominee_2_Name,
                                    "Nominee_2_Relationship": Nominee_2_Relationship,
                                    "Nominee_2_Applicable": Nominee_2_Applicable,
                                    "Nominee_2_DOB": Nominee_2_DOB,
                                    "Nominee_2_Minor_Flag": Nominee_2_Minor_Flag,
                                    "Nominee_2_Guardian": Nominee_2_Guardian,
                                    "Nominee_3_Name": Nominee_3_Name,
                                    "Nominee_3_Relationship": Nominee_3_Relationship,
                                    "Nominee_3_Applicable": Nominee_3_Applicable,
                                    "Nominee_3_DOB": Nominee_3_DOB,
                                    "Nominee_3_Minor_Flag": Nominee_3_Minor_Flag,
                                    "Nominee_3_Guardian": Nominee_3_Guardian,
                                    "Primary_Holder_KYC_Type": Primary_Holder_KYC_Type,
                                    "Primary_Holder_CKYC_Number": Primary_Holder_CKYC_Number,
                                    "Second_Holder_KYC_Type": Second_Holder_KYC_Type,
                                    "Second_Holder_CKYC_Number": Second_Holder_CKYC_Number,
                                    "Third_Holder_KYC_Type": Third_Holder_KYC_Type,
                                    "Third_Holder_CKYC_Number": Third_Holder_CKYC_Number,
                                    "Guardian_KYC_Type": Guardian_KYC_Type,
                                    "Guardian_CKYC_Number": Guardian_CKYC_Number,
                                    "Primary_Holder_KRA_Exempt_Ref_No": Primary_Holder_KRA_Exempt_Ref_No,
                                    "Second_Holder_KRA_Exempt_Ref_No": Second_Holder_KRA_Exempt_Ref_No,
                                    "Third_Holder_KRA_Exempt_Ref_No": Third_Holder_KRA_Exempt_Ref_No,
                                    "Guardian_Exempt_Ref_No": Guardian_Exempt_Ref_No,
                                    "Aadhaar_Updated": Aadhaar_Updated,
                                    "Mapin_Id": Mapin_Id,
                                    "Paperless_Flag": Paperless_Flag,
                                    "LEI_No": LEI_No,
                                    "LEI_Validity": LEI_Validity,
                                    "Filler_1_Mobile_Declaration_Flag": Filler_1_Mobile_Declaration_Flag,
                                    "Filler_2_Email_Declaration_Flag": Filler_2_Email_Declaration_Flag,
                                    "Nomination_Opt": Nomination_Opt,
                                    "Nomination_Auth_Mode": Nomination_Auth_Mode,
                                    "Nominee_PAN1": Nominee_PAN1,
                                    "Nominee_Guardian_PAN1": Nominee_Guardian_PAN1,
                                    "Nominee_PAN2": Nominee_PAN2,
                                    "Nominee_Guardian_PAN2": Nominee_Guardian_PAN2,
                                    "Nominee_PAN3": Nominee_PAN3,
                                    "Nominee_Guardian_PAN3": Nominee_Guardian_PAN3,
                                    "Second_Holder_Email": Second_Holder_Email,
                                    "Second_Holder_Email_Declaration": Second_Holder_Email_Declaration,
                                    "Second_Holder_Mobile_No": Second_Holder_Mobile_No,
                                    "Second_Holder_Mobile_No_Declaration": Second_Holder_Mobile_No_Declaration,
                                    "Third_Holder_Email": Third_Holder_Email,
                                    "Third_Holder_Email_Declaration": Third_Holder_Email_Declaration,
                                    "Third_Holder_Mobile_No": Third_Holder_Mobile_No,
                                    "Third_Holder_Mobile_No_Declaration": Third_Holder_Mobile_No_Declaration,
                                    "guardian_relationship": Guardian_Relationship,
                                }
                
                print(json.dumps(data['param'], indent=4))

            if id:
                try:
                    user = Uccregister.objects.get(id=id)
                except Uccregister.DoesNotExist:
                    return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

                serializer = UccregisterSerializer(user, data=data, context={'request': request}, partial=True)
                action = "updated"
            else:
                serializer = UccregisterSerializer(data=data, context={'request': request})
                action = "created"

            if serializer.is_valid(raise_exception=True):
                ucc = serializer.save()
                user = ucc.user 
                refresh = RefreshToken.for_user(user)
                access_token = {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }

                try:
                    response_data['status'] = True
                    response_data['message'] = f'User {action} successfully.'
                    response_data['data'] = serializer.data
                    response_data['data']['access_token'] = access_token
                    return Response(response_data, status=status.HTTP_201_CREATED if action == "created" else status.HTTP_200_OK)
                except Exception as e:
                    response_data['status'] = False
                    response_data['message'] = f'Error while {action} user: {str(e)}'
                    return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                response_data['status'] = False
                response_data['message'] = 'Error creating user'
                response_data['errors'] = serializer.errors
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        else:
            response_data['status'] = False
            response_data['message'] = "Invalid request method."
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
class FatcaAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        data = request.data

        if request.method == 'POST':
            user_id =data.get('user_id')
            pan_number = data.get('pan_number', "")
            pekrn = data.get('PEKRN', "")
            invester_name = data.get('invester_name', "")
            dob = data.get('dob', "")
            first_name = data.get('first_name', "")
            last_name = data.get('last_name', "")
            tax_status = data.get('tax_status', "")
            data_src = data.get('data_src', "")
            address_type = data.get('address_type', "")
            applications_type = data.get('applications_type', "")
            co_bir_inc = data.get('CO_BIR_INC', "")
            tax_res1 = data.get('TAX_RES1', "")
            tpin1 = data.get('tpin1', "")
            id1_type = data.get('ID1_TYPE', "")
            tax_res2 = data.get('TAX_RES2', "")
            tpin2 = data.get('TPIN2', "")
            id2_type = data.get('ID2_TYPE', "")
            tax_res3 = data.get('TAX_RES3', "")
            tpin3 = data.get('TPIN3', "")
            id3_type = data.get('ID3_TYPE', "")
            tax_res4 = data.get('TAX_RES4', "")
            tpin4 = data.get('TPIN4', "")
            id4_type = data.get('ID4_TYPE', "")
            srce_wealth = data.get('srce_wealth', "")
            corp_servs = data.get('CORP_SERVS', "")
            inc_slab = data.get('INC_SLAB', "")
            net_worth = data.get('NET_WORTH', "")
            nw_date = data.get('NW_DATE', "")
            pep_flag = data.get('pep_flag', "")
            occupation_code = data.get('occupation_code', "")
            occupation_Type = data.get('occupation_Type', "")
            exemp_code = data.get('exemp_code', "")
            ffi_drnfe = data.get('FFI_DRNFE', "")
            giin_no = data.get('GIIN_NO', "")
            spr_entity = data.get('SPR_ENTITY', "")
            giin_na = data.get('GIIN_NA', "")
            giin_exemc = data.get('GIIN_EXEMC', "")
            nffe_catg = data.get('NFFE_CATG', "")
            act_nfe_sc = data.get('ACT_NFE_SC', "")
            nature_bus = data.get('NATURE_BUS', "")
            rel_listed = data.get('REL_LISTED', "")
            exch_name = data.get('exch_name', "")
            ubo_appl = data.get('ubo_appl', "")
            ubo_count = data.get('ubo_count', "")
            ubo_name = data.get('ubo_name', "")
            ubo_pan = data.get('ubo_pan', "")
            ubo_nation = data.get('UBO_NATION', "")
            ubo_add1 = data.get('UBO_ADD1', "")
            ubo_add2 = data.get('UBO_ADD2', "")
            ubo_add3 = data.get('UBO_ADD3', "")
            ubo_city = data.get('UBO_CITY', "")
            ubo_pin = data.get('UBO_PIN', "")
            ubo_state = data.get('UBO_STATE', "")
            ubo_cntry = data.get('UBO_CNTRY', "")
            ubo_address_type = data.get('ubo_address_type', "")
            ubo_ctr = data.get('UBO_CTR', "")
            ubo_tin = data.get('UBO_TIN', "")
            ubo_id_ty = data.get('UBO_ID_TY', "")
            ubo_cob = data.get('UBO_COB', "")
            ubo_dob = data.get('ubo_dob', "")
            ubo_gender = data.get('ubo_gender', "")
            ubo_fr_nam = data.get('ubo_fr_nam', "")
            ubo_occ = data.get('ubo_occ', "")
            ubo_occ_ty = data.get('ubo_occ_ty', "")
            ubo_tel = data.get('UBO_TEL', "")
            ubo_mobile = data.get('UBO_MOBILE', "")
            ubo_code = data.get('ubo_code', "")
            ubo_hol_pc = data.get('UBO_HOL_PC', "")
            sdf_flag = data.get('SDF_FLAG', "")
            ubo_df = data.get('UBO_DF', "")
            aadhaar_rp = data.get('aadhaar_rp', "")
            new_change = data.get('new_change', "")
            log_name = data.get('log_name', "")
            ubo_exch = data.get('UBO_EXCH',"")
            ubo_isin = data.get('UBO_ISIN',"")
            filler1 = data.get('FILLER1', "")
            filler2 = data.get('FILLER2', "")

            try:
                date = datetime.strptime(dob, "%Y-%m-%d").strftime("%m/%d/%Y")

            except ValueError as e:
                response_data = {
                    "status": False,
                    "message": f"Enter a valid date format (YYYY-MM-DD). Error: {str(e)}"
                }
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
            agent_id = request.data.get('agent')

            if not agent_id:
                return Response({"detail": "Agent ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                agent = Agent.objects.get(agent_id=agent_id)
            except Agent.DoesNotExist:
                return Response({"detail": "Agent not found."}, status=status.HTTP_404_NOT_FOUND)

            request.data['agent'] = agent.id


            serializer = FATCASerializer(data=request.data)

            if not user_id:
                response_data['status'] = False
                response_data['message'] = "User ID is required"
                response_data['data'] = []
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                response_data["status"]= False
                response_data["message"]= "User not found."
                return Response(response_data, status=status.HTTP_404_NOT_FOUND)

            if user_id:
                data['user'] = user.id
            if pan_number:
                data['pan_number'] = pan_number
            if tax_status:
                data['tax_status'] = tax_status
            if invester_name:
                data['invester_name'] = invester_name
            if dob:
                data['dob'] = dob
            if data_src:
                data['data_src'] = data_src
            if address_type:
                data['address_type'] = address_type
            if applications_type:
                data['applications_type'] = applications_type
            if tpin1:
                data['tpin1'] = tpin1
            if srce_wealth:
                data['srce_wealt'] = srce_wealth
            if pep_flag:
                data['pep_flag'] = pep_flag
            if occupation_code:
                data['occupation_code'] = occupation_code
            if occupation_Type:
                data['occupation_Type'] = occupation_Type
            if exch_name:
                data['exch_name'] = exch_name
            if ubo_appl:
                data['ubo_appl'] = ubo_appl
            if ubo_count:
                data['ubo_count'] = ubo_count
            if ubo_name:
                data['ubo_name'] = ubo_name
            if ubo_pan:
                data['ubo_pan'] = ubo_pan
            if ubo_address_type:
                data['ubo_address_type'] = ubo_address_type
            if ubo_gender:
                data['ubo_gender'] = ubo_gender
            if ubo_fr_nam:
                data['ubo_fr_nam'] = ubo_fr_nam
            if ubo_occ:
                data['ubo_occ'] = ubo_occ
            if ubo_occ_ty:
                data['ubo_occ_ty'] = ubo_occ_ty
            if ubo_code:
                data['ubo_code'] = ubo_code
            if aadhaar_rp:
                data['aadhaar_rp'] = aadhaar_rp
            if new_change:
                data['new_change'] = new_change
            if log_name:
                data['log_name'] = log_name
            if ubo_exch:
                data['UBO_EXCH'] = ubo_exch

            encrypted_password = get_password()

            soap_message = f"""
            <soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" xmlns:ns="http://bsestarmfdemo.bseindia.com/2016/01/">
                <soap:Header xmlns:wsa="http://www.w3.org/2005/08/addressing">
                    <wsa:Action>http://bsestarmfdemo.bseindia.com/2016/01/IMFUploadService/MFAPI</wsa:Action>
                    <wsa:To>https://bsestarmfdemo.bseindia.com/MFUploadService/MFUploadService.svc/Secure</wsa:To>
                </soap:Header>
                <soap:Body>
                    <ns:MFAPI>
                        <ns:Flag>01</ns:Flag>
                        <ns:UserId>{loginid}</ns:UserId>
                        <ns:EncryptedPassword>{encrypted_password}</ns:EncryptedPassword>
                        <ns:param>
                            {pan_number}|{pekrn}|{invester_name}|{date}|{first_name}|{last_name}|{tax_status}|{data_src}|{address_type}|{applications_type}|{co_bir_inc}|{tax_res1}|{tpin1}|{id1_type}|{tax_res2}|{tpin2}|{id2_type}|{tax_res3}|{tpin3}|{id3_type}|{tax_res4}|{tpin4}|{id4_type}|{srce_wealth}|{corp_servs}|{inc_slab}|{net_worth}|{nw_date}|{pep_flag}|{occupation_code}|{occupation_Type}|{exemp_code}|{ffi_drnfe}|{giin_no}|{spr_entity}|{giin_na}|{giin_exemc}|{nffe_catg}|{act_nfe_sc}|{nature_bus}|{rel_listed}|{exch_name}|{ubo_appl}|{ubo_count}|{ubo_name}|{ubo_pan}|{ubo_nation}|{ubo_add1}|{ubo_add2}|{ubo_add3}|{ubo_city}|{ubo_pin}|{ubo_state}|{ubo_cntry}|{ubo_address_type}|{ubo_ctr}|{ubo_tin}|{ubo_id_ty}|{ubo_cob}|{ubo_dob}|{ubo_gender}|{ubo_fr_nam}|{ubo_occ}|{ubo_occ_ty}|{ubo_tel}|{ubo_mobile}|{ubo_code}|{ubo_hol_pc}|{sdf_flag}|{ubo_df}|{aadhaar_rp}|{new_change}|{log_name}|{ubo_exch}|{ubo_isin}
                        </ns:param>
                    </ns:MFAPI>
                </soap:Body>
            </soap:Envelope>
            """
            conn = http.client.HTTPSConnection(base_url)

            headers = {
                'Content-Type': 'application/soap+xml; charset=utf-8',
                'SOAPAction': 'http://webservice.bseindia.com/SubmitRequest',
            }

            conn.request("POST", "/MFUploadService/MFUploadService.svc/Secure", soap_message, headers)
            res = conn.getresponse()
            api_data = res.read().decode()
            print(res.status, res.reason, api_data, 'jjjjjjjjjjjj')

            if res.status == 503:
                response_data['status'] = False
                response_data['message'] = "Service Unavailable. Please try again later."
                return Response(response_data, status=503)

            try:
                api_data_dict = xmltodict.parse(api_data)
                api_data_json = json.dumps(api_data_dict)
            except Exception as e:
                response_data['status'] = False
                response_data['message'] = f"Error parsing XML response: {str(e)}"
                return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            data_dict = json.loads(api_data_json)

            print("Parsed XML as JSON:", data_dict)

            mfapi_result = data_dict.get('s:Envelope', {}).get('s:Body', {}).get('MFAPIResponse', {}).get('MFAPIResult', '')

            print("MFAPIResult:", mfapi_result)

            result_parts = mfapi_result.split('|')
            status_code = result_parts[0].strip() if len(result_parts) > 0 else ""
            status_message = result_parts[1].strip() if len(result_parts) > 1 else "Unknown error"

            if status_code != '100':
                response_data['status'] = False
                response_data['message'] = f"Error: {status_message}"
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            data['param']= {
                    "pan_number": pan_number,
                    "pekrn": pekrn,
                    "invester_name": invester_name,
                    "dob": dob,
                    "first_name": first_name,
                    "last_name": last_name,
                    "tax_status": tax_status,
                    "data_src": data_src,
                    "address_type": address_type,
                    "applications_type": applications_type,
                    "co_bir_inc": co_bir_inc,
                    "tax_res1": tax_res1,
                    "tpin1": tpin1,
                    "id1_type": id1_type,
                    "tax_res2": tax_res2,
                    "tpin2": tpin2,
                    "id2_type": id2_type,
                    "tax_res3": tax_res3,
                    "tpin3": tpin3,
                    "id3_type": id3_type,
                    "tax_res4": tax_res4,
                    "tpin4": tpin4,
                    "id4_type": id4_type,
                    "srce_wealth": srce_wealth,
                    "corp_servs": corp_servs,
                    "inc_slab": inc_slab,
                    "net_worth": net_worth,
                    "nw_date": nw_date,
                    "pep_flag": pep_flag,
                    "occupation_code": occupation_code,
                    "occupation_Type": occupation_Type,
                    "exemp_code": exemp_code,
                    "ffi_drnfe": ffi_drnfe,
                    "giin_no": giin_no,
                    "spr_entity": spr_entity,
                    "giin_na": giin_na,
                    "giin_exemc": giin_exemc,
                    "nffe_catg": nffe_catg,
                    "act_nfe_sc": act_nfe_sc,
                    "nature_bus": nature_bus,
                    "rel_listed": rel_listed,
                    "exch_name": exch_name,
                    "ubo_appl": ubo_appl,
                    "ubo_count": ubo_count,
                    "ubo_name": ubo_name,
                    "ubo_pan": ubo_pan,
                    "ubo_nation": ubo_nation,
                    "ubo_add1": ubo_add1,
                    "ubo_add2": ubo_add2,
                    "ubo_add3": ubo_add3,
                    "ubo_city": ubo_city,
                    "ubo_pin": ubo_pin,
                    "ubo_state": ubo_state,
                    "ubo_cntry": ubo_cntry,
                    "ubo_address_type": ubo_address_type,
                    "ubo_ctr": ubo_ctr,
                    "ubo_tin": ubo_tin,
                    "ubo_id_ty": ubo_id_ty,
                    "ubo_cob": ubo_cob,
                    "ubo_dob": ubo_dob,
                    "ubo_gender": ubo_gender,
                    "ubo_fr_nam": ubo_fr_nam,
                    "ubo_occ": ubo_occ,
                    "ubo_occ_ty": ubo_occ_ty,
                    "ubo_tel": ubo_tel,
                    "ubo_mobile": ubo_mobile,
                    "ubo_code": ubo_code,
                    "ubo_hol_pc": ubo_hol_pc,
                    "sdf_flag": sdf_flag,
                    "ubo_df": ubo_df,
                    "aadhaar_rp": aadhaar_rp,
                    "new_change": new_change,
                    "log_name": log_name,
                    "ubo_exch": ubo_exch,
                    "ubo_isin": ubo_isin,
                    "filler1": filler1,
                    "filler2": filler2
                }

            data['api_data'] = data_dict

            if new_change == 'C':
                try:
                    existing_record = Fatca.objects.get(pan_number=pan_number)
                    for field, value in data.items():
                        setattr(existing_record, field, value)
                    existing_record.save()

                    response_data['status'] = True
                    response_data['message'] = 'FATCA record updated successfully'
                    response_data['data'] = {'id': existing_record.id, **data}
                    return Response(response_data, status=status.HTTP_404_NOT_FOUND)
                except Fatca.DoesNotExist:
                    response_data['status'] = False
                    response_data['message'] = 'Record not found for updating'
                    return Response(response_data, status=status.HTTP_404_NOT_FOUND)
                
            if Fatca.objects.filter(pan_number=pan_number).exists():
                response_data['status'] = False
                response_data['message'] = "FATCA record with this PAN number already exists."
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
                
            serializer = FATCASerializer(data=data)
            if serializer.is_valid(raise_exception=True):
                serializer.save()

                try:
                    response_data['status'] = True
                    response_data['message'] = 'FATCA Registration Done Successfully'
                    response_data['data'] = serializer.data
                    return Response(response_data, status=status.HTTP_201_CREATED)
                except Exception as e:
                    response_data['status'] = False
                    response_data['message'] = f'Error sending: {str(e)}'
                    return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                response_data['status'] = False
                response_data['message'] = 'FATCA Registration failed'
                response_data['errors'] = serializer.errors
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
class MandateAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        data = request.data

        user_id  = data.get('user_id')
        client_code = data.get('client_code')
        amount = data.get('amount')
        mandate_type = data.get('mandate_type')
        account_no = data.get('account_no')
        account_type = data.get('account_type')
        ifsc_code = data.get('ifsc_code')
        misc_code = data.get('misc_code', "")
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if not user_id:
                response_data['status'] = False
                response_data['message'] = "User ID is required"
                response_data['data'] = []
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data["status"]= False
            response_data["message"]= "User not found."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        required_fields = [client_code, amount, mandate_type, account_no, account_type, ifsc_code, start_date, end_date]
        if not all(required_fields):
            response_data['status'] = False
            response_data['message'] = "Missing required fields"
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            formatted_start_date = datetime.strptime(start_date, "%Y-%m-%d").strftime("%d/%m/%Y")
            formatted_end_date = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError as e:
            response_data = {
                "status": False,
                "message": f"Enter a valid date format (YYYY-MM-DD). Error: {str(e)}"
            }
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        agent_id = request.data.get('agent')

        if not agent_id:
            return Response({"detail": "Agent ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(agent_id=agent_id)
        except Agent.DoesNotExist:
            return Response({"detail": "Agent not found."}, status=status.HTTP_404_NOT_FOUND)

        request.data['agent'] = agent.id
    
        encrypted_password = get_password()

        soap_message = f"""
        <soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" xmlns:ns="http://bsestarmfdemo.bseindia.com/2016/01/">
            <soap:Header xmlns:wsa="http://www.w3.org/2005/08/addressing">
                <wsa:Action>http://bsestarmfdemo.bseindia.com/2016/01/IMFUploadService/MFAPI</wsa:Action>
                <wsa:To>https://bsestarmfdemo.bseindia.com/MFUploadService/MFUploadService.svc/Secure</wsa:To>
            </soap:Header>
            <soap:Body>
                <ns:MFAPI>
                    <ns:Flag>06</ns:Flag>
                    <ns:UserId>{loginid}</ns:UserId>
                    <ns:EncryptedPassword>{encrypted_password}</ns:EncryptedPassword>
                    <ns:param>{client_code}|{amount}|{mandate_type}|{account_no}|{account_type}|{ifsc_code}|{misc_code}|{formatted_start_date}|{formatted_end_date}</ns:param>
                </ns:MFAPI>
            </soap:Body>
        </soap:Envelope>
        """

        try:
            conn = http.client.HTTPSConnection(base_url)
            headers = {
                'Content-Type': 'application/soap+xml; charset=utf-8',
                'SOAPAction': 'http://webservice.bseindia.com/SubmitRequest',
            }
            conn.request("POST", "/MFUploadService/MFUploadService.svc/Secure", soap_message, headers)
            res = conn.getresponse()
            api_data = res.read().decode()
            print(api_data, 'api data')
        except Exception as e:
            response_data['status'] = False
            response_data['message'] = f"SOAP Request Failed: {str(e)}"
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        if res.status == 503:
                response_data['status'] = False
                response_data['message'] = "Service Unavailable. Please try again later."
                return Response(response_data, status=503)
        
        try:
            api_data_dict = xmltodict.parse(api_data)
            data_dict = json.loads(json.dumps(api_data_dict))
        except Exception as e:
            response_data['status'] = False
            response_data['message'] = f"Error parsing XML response: {str(e)}"
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        data['api_data']= data_dict
        data['user'] = user.id

        mfapi_result = data_dict.get('s:Envelope', {}).get('s:Body', {}).get('MFAPIResponse', {}).get('MFAPIResult', '')

        print("MFAPIResult:", mfapi_result)

        if "MFAPIResult:" in mfapi_result:
            mfapi_result = mfapi_result.split("MFAPIResult:")[-1].strip()

        result_parts = mfapi_result.split('|')

        status_code = result_parts[0].strip() if len(result_parts) > 0 else ''
        status_message = result_parts[1].strip() if len(result_parts) > 1 else 'Unknown error'

        print(f"Status Code: {status_code}")
        print(f"Status Message: {status_message}")

        if status_code == '100':
            try:
                if len(result_parts) >= 3:
                    mandate_id = result_parts[2].strip()
                    data['mandate_id'] = mandate_id
                else:
                    response_data['status'] = False
                    response_data['message'] = 'Mandate ID not found in success message'
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

                serializer = MandateSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    response_data['status'] = True
                    response_data['message'] = 'Mandate Registration Done Successfully'
                    response_data['data'] = serializer.data
                    return Response(response_data, status=status.HTTP_201_CREATED)
                else:
                    response_data['status'] = False
                    response_data['message'] = 'Mandate Registration failed'
                    response_data['errors'] = serializer.errors
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                response_data['status'] = False
                response_data['message'] = f"Error extracting Mandate ID: {str(e)}"
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        else:
            response_data['status'] = False
            response_data['message'] = f"Error: {status_message}"
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
class EnachAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        data = request.data

        user_id = data.get('user_id')
        client_code = data.get('client_code')
        mandate_id = data.get('mandate_id')

        if not user_id:
                response_data['status'] = False
                response_data['message'] = "User ID is required"
                response_data['data'] = []
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST) 
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data["status"]= False
            response_data["message"]= "User not found."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)
        
        agent_id = request.data.get('agent')

        if not agent_id:
            return Response({"detail": "Agent ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(agent_id=agent_id)
        except Agent.DoesNotExist:
            return Response({"detail": "Agent not found."}, status=status.HTTP_404_NOT_FOUND)

        request.data['agent'] = agent.id

        payload = json.dumps({
            "UserId": loginid,
            "MemberCode":membercode,
            "Password": password,
            "ClientCode": client_code,
            "MandateID": mandate_id
        })
        print(payload, 'Payload Sent to API')

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        try:
            conn = http.client.HTTPSConnection(base_url)
            conn.request("POST", "/StarMFWebService/StarMFWebService.svc/EMandateAuthURL", payload, headers)
            res = conn.getresponse()
            api_data = res.read().decode()
            content_type = res.getheader('Content-Type', '')

            if 'text/html' in content_type:
                response_data['status'] = False
                response_data['message'] = 'API returned an HTML error page. Likely due to invalid payload or missing headers.'
                response_data['errors'] = api_data
                return Response(response_data, status=status.HTTP_502_BAD_GATEWAY)

            print(content_type, "Content-Type of Response")
            print(api_data, 'Raw API Response')

        except Exception as e:
            response_data['status'] = False
            response_data['message'] = 'API Request Failed'
            response_data['errors'] = str(e)
            return Response(response_data, status=status.HTTP_502_BAD_GATEWAY)

        try:
            data_dict = json.loads(api_data)
            print(data_dict, 'JSON Parsed Successfully')
        except json.JSONDecodeError:
            try:
                api_data_dict = xmltodict.parse(api_data)
                data_dict = json.loads(json.dumps(api_data_dict))
                print(data_dict, 'XML Parsed Successfully')
            except Exception as e:
                response_data['status'] = False
                response_data['message'] = 'Error parsing API response'
                response_data['errors'] = str(e)
                return Response(response_data, status=status.HTTP_502_BAD_GATEWAY)
            
        api_status = data_dict.get('Status', '')
        if api_status != '100':
            error_message = data_dict.get('ResponseString', 'Unknown error')
            return Response({
                'status': False,
                'message': 'E-NACH Registration Failed',
                'errors': error_message
            }, status=status.HTTP_400_BAD_REQUEST)

        data['api_data'] = data_dict
        data['user'] = user.id 

        serializer = EnachSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            response_data['status'] = True
            response_data['message'] = 'E-NACH Registration Done Successfully'
            response_data['data'] = serializer.data
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            response_data['status'] = False
            response_data['message'] = 'E-NACH Registration failed'
            response_data['errors'] = serializer.errors
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
class XSIPAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        response_data = {}
        data = request.data

        user_id = data.get("user_id")
        scheme_code = data.get("scheme_code")
        client_code = data.get("client_code")
        int_ref_no = data.get("int_ref_no","")
        trans_mode = data.get("trans_mode")
        dp_trans_mode = data.get("dp_trans_mode")
        start_date = data.get("start_date")
        frequency_type = data.get("frequency_type")
        frequency_allowed = data.get("frequency_allowed")
        installments_amount = data.get("installments_amount")
        no_of_installments = data.get("no_of_installments")
        remarks = data.get("remarks","")
        folio_no = data.get("folio_no","")
        first_order_flag = data.get("first_order_flag")
        sub_br_code = data.get("sub_br_code","")
        euin = data.get("euin","")
        euin_flag = data.get("euin_flag")
        dpc = data.get("dpc")
        sub_broker_arn = data.get("sub_broker_arn","")
        end_date = data.get("EndDate") or None 
        regn_type = data.get("regn_type")
        brokerage = data.get("brokerage","")
        mandate_id = data.get("mandate_id")
        xsip_type = data.get("xsip_type") 
        target_scheme = data.get("target_scheme","")
        target_amount = data.get("target_amount","")
        goal_type = data.get("goal_type","") 
        goal_amount = data.get("goal_amount","")
        filler1 = data.get("filler1","") 
        filler2 = data.get("filler2","") 
        filler3 = data.get("filler3","") 
        filler4 = data.get("filler4","") 
        filler5 = data.get("filler5","") 
        filler6 = data.get("filler6","") 
        filler7 = data.get("filler7","") 
        filler8 = data.get("filler8","") 
        filler9 = data.get("filler9","") 
        filler10 = data.get("filler10","")

        try:
            formatted_start_date = datetime.strptime(start_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError as e:
            response_data = {
                "status": False,
                "message": f"Enter a valid date format (YYYY-MM-DD). Error: {str(e)}"
            }
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        if not user_id:
            response_data['status'] = False
            response_data['message'] = "User ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data["status"]= False
            response_data["message"]= "User not found."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        data['user']=user.id

        agent_id = request.data.get('agent')

        if not agent_id:
            return Response({"detail": "Agent ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(agent_id=agent_id)
        except Agent.DoesNotExist:
            return Response({"detail": "Agent not found."}, status=status.HTTP_404_NOT_FOUND)

        request.data['agent'] = agent.id

        payload = json.dumps({
            "LoginId" : loginid,
            "MemberCode" : membercode,
            "Password" : password,
            "SchemeCode" : scheme_code,
            "ClientCode" : client_code,
            "IntRefNo" : int_ref_no,
            "TransMode" : trans_mode,
            "DPTransMode" : dp_trans_mode,
            "StartDate" : formatted_start_date,
            "FrequencyType" : frequency_type,
            "FrequencyAllowed" : frequency_allowed,
            "InstAmount" : installments_amount,
            "NoOfInst" : no_of_installments,
            "Remarks" : remarks,
            "FolioNo" : folio_no,
            "FirstOrderFlag" : first_order_flag,
            "SubBrCode" : sub_br_code,
            "EUIN" : euin,
            "EUINFlag" : euin_flag,
            "DPC" : dpc,
            "SubBrokerARN" : sub_broker_arn, 
            "EndDate" :  end_date,
            "RegnType" : regn_type,
            "Brokerage" : brokerage,
            "MandateId" : mandate_id,
            "XSIPType" :  xsip_type,
            "TargetScheme" : target_scheme,
            "TargetAmount" : target_amount,
            "GoalType" :  goal_type,
            "GoalAmount" : goal_amount,
            "Filler1" :  filler1,
            "Filler2" :  filler2,
            "Filler3" :  filler3,
            "Filler4" :  filler4,
            "Filler5" :  filler5,
            "Filler6" :  filler6,
            "Filler7" :  filler7,
            "Filler8" :  filler8,
            "Filler9" :  filler9,
            "Filler10" : filler10
        })
        print(payload,"ppppppppppppppppppppppppppppppppp")

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'APIKEY': 'VmxST1UyRkhUbkpOVldNOQ=='
        }

        try:
            conn = http.client.HTTPSConnection(base_url)
            conn.request("POST", "/StarMFAPI/api/XSIP/XSIPRegistration", payload, headers)
            res = conn.getresponse()
            api_data = res.read().decode()

            print(f"API Response Status: {res.status}")
            print(f"API Response Data: {api_data}")
        except Exception as e:
            print(f"API Request Failed: {str(e)}")


        try:
            data_dict = json.loads(api_data)
            print(data_dict, 'JSON Parsed Successfully')
        except json.JSONDecodeError:
            try:
                api_data_dict = xmltodict.parse(api_data)
                data_dict = json.loads(json.dumps(api_data_dict))
                print(data_dict, 'XML Parsed Successfully')
            except Exception as e:
                response_data['status'] = False
                response_data['message'] = 'Error parsing API response'
                response_data['errors'] = str(e)
                return Response(response_data, status=status.HTTP_502_BAD_GATEWAY)
            
        api_xsipregid = data_dict.get('XSIPRegId', None)

        data['xsipregid'] = api_xsipregid

        api_status = data_dict.get('SuccessFlag', '')
        if api_status != '0':
            error_message = data_dict.get('BSERemarks') or data_dict.get('ResponseString') or "Unknown error"

            print(f"API Error Message: {error_message}") 
            return Response({
                'status': False,
                'message': 'XSIP Registration Failed',
                'errors': error_message
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = XSIPSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            response_data['status'] = True
            response_data['message'] = 'XSIP Registration Done Successfully'
            response_data['data'] = serializer.data
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            response_data['status'] = False
            response_data['message'] = 'XSIP Registration failed'
            response_data['errors'] = serializer.errors
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request):
        response_data = {}
        data = request.data

        user_id = data.get("user_id")
        agent_id = request.data.get('agent')
        client_code = data.get("client_code")
        xsipregid = data.get("RegnNo")
        cease_bse_code = data.get("CeaseBseCode","")
        int_ref_no = data.get("int_ref_no","")
        remarks = data.get("remarks","")

        if not agent_id:
            response_data['status'] = False
            response_data['message'] = "Agent ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            agent = Agent.objects.get(agent_id=agent_id)
        except Agent.DoesNotExist:
            response_data["status"]= False
            response_data["message"]= "Agent not found."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)
        
        if not user_id:
            response_data['status'] = False
            response_data['message'] = "User ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data["status"]= False
            response_data["message"]= "User not found."
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)
        
        if not client_code:
            response_data['status'] = False
            response_data['message'] = "Client code is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        if not xsipregid:
            response_data['status'] = False
            response_data['message'] = "XSIP Registration. ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        if not cease_bse_code:
            response_data['status'] = False
            response_data['message'] = "BSE code is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        payload = json.dumps({
            "LoginId" : loginid,
            "MemberCode" : membercode,
            "Password" : password,
            "ClientCode" : client_code,
            "RegnNo" : xsipregid,
            "CeaseBseCode" :cease_bse_code,
            "IntRefNo" : int_ref_no,
            "Remarks" : remarks,
        })
        print(payload,"payload")

        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'APIKEY': 'VmxST1UyRkhUbkpOVldNOQ=='
        }

        try:
            conn = http.client.HTTPSConnection(base_url)
            conn.request("POST", "/StarMFAPI/api/XSIP/XSIPCancellation", payload, headers)
            res = conn.getresponse()
            api_data = res.read().decode()

            print(f"API Response Status: {res.status}")
            print(f"API Response Data: {api_data}")
        except Exception as e:
            print(f"API Request Failed: {str(e)}")

        try:
            data_dict = json.loads(api_data)
            print(data_dict, 'JSON Parsed Successfully')
        except json.JSONDecodeError:
            try:
                api_data_dict = xmltodict.parse(api_data)
                data_dict = json.loads(json.dumps(api_data_dict))
                print(data_dict, 'XML Parsed Successfully')
            except Exception as e:
                response_data['status'] = False
                response_data['message'] = 'Error parsing API response'
                response_data['errors'] = str(e)
                return Response(response_data, status=status.HTTP_502_BAD_GATEWAY)
            
        api_status = data_dict.get('SuccessFlag', '')
        if api_status != '0':
            error_message = data_dict.get('BSERemarks') or data_dict.get('ResponseString') or "Unknown error"

            print(f"API Error Message: {error_message}") 
            return Response({
                'status': False,
                'message': 'XSIP Cancellation  Failed',
                'errors': error_message
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = XSIPSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            response_data['status'] = True
            response_data['message'] = 'XSIP Cancellation Successfully'
            response_data['data'] = serializer.data
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            response_data['status'] = False
            response_data['message'] = 'XSIP Cancellation failed'
            response_data['errors'] = serializer.errors
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

class AMCAPIView(APIView):
    def get(self, request):
        try:
            all_amc = AMCList.objects.all().order_by("-created_at")
            if not all_amc.exists():
                return Response({"message": "No AMC data found"}, status=status.HTTP_404_NOT_FOUND)

            latest = all_amc.first()
            param_data = latest.param

            filtered_data = []
            for row in param_data:
                scheme_code = row.get("Scheme Code")
                scheme_name = row.get("Scheme Name")

                if scheme_code or scheme_name:
                    filtered_data.append({
                        "Scheme Code": scheme_code,
                        "Scheme Name": scheme_name
                    })

            return Response({"data": filtered_data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def post(self, request):
        if not request.user.is_superuser:
         return Response({"message": "Only Admin can this document."}, status=status.HTTP_403_FORBIDDEN)
        
        uploaded_file = request.FILES.get("AMC_list_file")

        if not uploaded_file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.endswith(".xlsx"):
            return Response({"error": "Only .xlsx files are supported"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            headers = [str(h).strip() if h else "" for h in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                clean_row = {}
                for key, value in zip(headers, row):
                    if isinstance(value, time):
                        value = value.strftime("%H:%M:%S")
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")
                    clean_row[key] = value

                data.append(clean_row)

            amc_instance = AMCList.objects.create(
                AMC_list_file=uploaded_file,
                param=data
            )

            serializer = AMCListSerializer(amc_instance)
            return Response({
                "message": f"{len(data)} rows uploaded and saved successfully.",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class AMCCartAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_id = request.query_params.get('user_id')
        response_data = {}

        if not user_id:
            response_data['status'] = False
            response_data['message'] = 'User ID is required'
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data['status'] = False
            response_data['message'] = 'User not found'
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        cart_items = AMCCart.objects.filter(user=user)
        serializer = AMCCartSerializer(cart_items, many=True)
        return Response({'status': True, 'data': serializer.data}, status=status.HTTP_200_OK)
    
    def post(self, request):
        response_data = {}
        data = request.data

        user_id = data.get("user_id")
        scheme_code = data.get("scheme_code")
        amount = data.get("amount")
        start_date = data.get("start_date")

        if not user_id:
            response_data['status'] = False
            response_data['message'] = "User ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        if not scheme_code:
            response_data['status'] = False
            response_data['message'] = "Scheme Code is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        
        if not amount:
            response_data['status'] = False
            response_data['message'] = "Amount is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        if not start_date:
            response_data['status'] = False
            response_data['message'] = "Start Date is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data['status'] = False
            response_data['message'] = "User not found"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        amc_list = AMCList.objects.last()
        param_data = amc_list.param if amc_list else []

        matched = next((item for item in param_data if item.get("Scheme Code") == scheme_code), None)

        if not matched:
            response_data['status'] = False
            response_data['message'] = "Scheme Code not found in AMC List"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        scheme_name = matched.get("Scheme Name")

        if AMCCart.objects.filter(user=user, scheme_code=scheme_code).exists():
            response_data['status'] = False
            response_data['message'] = "This scheme is already in the cart"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_409_CONFLICT)

        serializer = AMCCartSerializer(data={
            "user": user.id,
            "scheme_code": scheme_code,
            "scheme_name": scheme_name,
            "amount": amount,
            "start_date": start_date
        })

        if serializer.is_valid():
            serializer.save()
            response_data['status'] = True
            response_data['message'] = "AMC scheme added to cart successfully"
            response_data['data'] = serializer.data
            return Response(response_data, status=status.HTTP_201_CREATED)

        response_data['status'] = False
        response_data['message'] = "Invalid data"
        response_data['data'] = serializer.errors
        return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request):
        response_data = {}
        user_id = request.query_params.get('user_id')
        scheme_code = request.query_params.get('scheme_code')


        if not user_id:
            response_data['status'] = False
            response_data['message'] = "User ID is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        if not scheme_code:
            response_data['status'] = False
            response_data['message'] = "Scheme Code is required"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            response_data['status'] = False
            response_data['message'] = "User not found"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        if request.user != user:
            response_data['status'] = False
            response_data['message'] = "You are not authorized to delete cart items for this user"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_403_FORBIDDEN)

        try:
            cart_item = AMCCart.objects.get(user=user, scheme_code=scheme_code)
        except AMCCart.DoesNotExist:
            response_data['status'] = False
            response_data['message'] = "Item not found in cart"
            response_data['data'] = []
            return Response(response_data, status=status.HTTP_404_NOT_FOUND)

        cart_item.delete()

        response_data['status'] = True
        response_data['message'] = "Item removed from cart successfully"
        response_data['data'] = []
        return Response(response_data, status=status.HTTP_200_OK)